"""
Seed script for OPC GPMB system.
Reads the Excel file and populates the database with:
  1. WorkflowTemplate + WorkflowNodes (from sheet 1)
  2. 4 Users
  3. HoSoGPMB (+ snapshot + tasks)
  4. 453 Ho (from sheet 2)
  5. Scope assignments + task instances for all per_household nodes
  6. Demo: complete C001-C010 for HB001

Run from backend directory:
  python -m scripts.seed
or directly:
  python scripts/seed.py
"""
from __future__ import annotations

import os
import sys
import uuid
from datetime import datetime
from typing import Optional, Dict, List

# Allow running from backend/ directory
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from passlib.context import CryptContext
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker, Session

# ── Config ────────────────────────────────────────────────────────────────────

DB_URL_SYNC = os.environ.get(
    "DATABASE_URL_SYNC",
    "postgresql+psycopg2://opc:opc_secret@localhost:5432/opc_gpmb",
)
EXCEL_PATH = os.environ.get(
    "EXCEL_PATH",
    "/Users/thanhld/Entrance/Project/ai-agent-practice/odin-gpmb/Quy trinh GPMB (1).xlsx",
)

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

# ── SQLAlchemy setup (sync) ───────────────────────────────────────────────────

from app.db.base import Base  # noqa: E402
from app.db.models import (  # noqa: E402
    User,
    RoleEnum,
    WorkflowTemplate,
    WorkflowNode,
    HoSoGPMB,
    HoSoStatusEnum,
    HoSoWorkflowNode,
    Ho,
    HoStatusEnum,
    NodeHouseholdScope,
    TaskInstance,
    TaskStatusEnum,
)

engine = create_engine(DB_URL_SYNC, echo=False)
SessionLocal = sessionmaker(bind=engine)


# ── Helpers ───────────────────────────────────────────────────────────────────

def create_wf_node(
    session: Session,
    template_id: uuid.UUID,
    name: str,
    level: int,
    order: int,
    code: Optional[str] = None,
    parent_id: Optional[uuid.UUID] = None,
    per_household: bool = False,
    planned_days: Optional[int] = None,
    legal_basis: Optional[str] = None,
    org_in_charge: Optional[str] = None,
    org_coordinate: Optional[str] = None,
) -> WorkflowNode:
    node = WorkflowNode(
        id=uuid.uuid4(),
        template_id=template_id,
        parent_id=parent_id,
        code=code,
        name=name,
        level=level,
        order=order,
        per_household=per_household,
        planned_days=planned_days,
        legal_basis=legal_basis,
        org_in_charge=org_in_charge,
        org_coordinate=org_coordinate,
    )
    session.add(node)
    session.flush()
    return node


def snapshot_workflow(
    session: Session, ho_so_id: uuid.UUID, template_id: uuid.UUID
) -> Dict[uuid.UUID, uuid.UUID]:
    """
    Snapshot WorkflowNodes → HoSoWorkflowNodes.
    Returns mapping: original_node_id → new_snapshot_node_id
    """
    nodes: List[WorkflowNode] = (
        session.query(WorkflowNode)
        .filter(WorkflowNode.template_id == template_id)
        .order_by(WorkflowNode.level, WorkflowNode.order)
        .all()
    )

    id_map: Dict[uuid.UUID, uuid.UUID] = {}

    # First pass: create without parent
    for node in nodes:
        new_id = uuid.uuid4()
        snap = HoSoWorkflowNode(
            id=new_id,
            ho_so_id=ho_so_id,
            source_node_id=node.id,
            parent_id=None,
            code=node.code,
            name=node.name,
            level=node.level,
            order=node.order,
            planned_days=node.planned_days,
            is_milestone=node.is_milestone,
            legal_basis=node.legal_basis,
            org_in_charge=node.org_in_charge,
            org_coordinate=node.org_coordinate,
            per_household=node.per_household,
            require_scan=node.require_scan,
            field_so_vb=node.field_so_vb,
            field_ngay_vb=node.field_ngay_vb,
            field_loai_vb=node.field_loai_vb,
            field_gia_tri_trinh=node.field_gia_tri_trinh,
            field_gia_tri_duyet=node.field_gia_tri_duyet,
            field_ghi_chu=node.field_ghi_chu,
        )
        session.add(snap)
        id_map[node.id] = new_id

    session.flush()

    # Second pass: set parent_id
    for node in nodes:
        if node.parent_id and node.parent_id in id_map:
            new_snap = session.get(HoSoWorkflowNode, id_map[node.id])
            new_snap.parent_id = id_map[node.parent_id]

    session.flush()
    return id_map


def generate_non_per_hh_tasks(session: Session, ho_so_id: uuid.UUID) -> int:
    """Create TaskInstance for all non-per_household snapshot nodes."""
    nodes: List[HoSoWorkflowNode] = (
        session.query(HoSoWorkflowNode)
        .filter(
            HoSoWorkflowNode.ho_so_id == ho_so_id,
            HoSoWorkflowNode.per_household == False,
        )
        .all()
    )
    count = 0
    for node in nodes:
        existing = (
            session.query(TaskInstance)
            .filter(
                TaskInstance.workflow_node_id == node.id,
                TaskInstance.ho_id == None,
            )
            .first()
        )
        if existing is None:
            task = TaskInstance(
                id=uuid.uuid4(),
                ho_so_id=ho_so_id,
                workflow_node_id=node.id,
                ho_id=None,
                status=TaskStatusEnum.dang_thuc_hien,
            )
            session.add(task)
            count += 1
    session.flush()
    return count


def assign_all_households_to_node(
    session: Session,
    ho_so_id: uuid.UUID,
    snap_node_id: uuid.UUID,
    ho_list: List[Ho],
) -> None:
    """
    Assign all households to a per_household snap node and all its per_household descendants.
    Creates NodeHouseholdScope + TaskInstance for each combination.
    """
    # Gather all per_household descendants (BFS)
    def get_descendants(node_id: uuid.UUID) -> List[HoSoWorkflowNode]:
        children = (
            session.query(HoSoWorkflowNode)
            .filter(
                HoSoWorkflowNode.ho_so_id == ho_so_id,
                HoSoWorkflowNode.parent_id == node_id,
            )
            .all()
        )
        result = list(children)
        for child in children:
            result.extend(get_descendants(child.id))
        return result

    root_node = session.get(HoSoWorkflowNode, snap_node_id)
    all_nodes = [root_node] + get_descendants(snap_node_id)
    per_hh_nodes = [n for n in all_nodes if n.per_household]

    print(f"  Assigning {len(ho_list)} households to node '{root_node.name[:40]}' ({len(per_hh_nodes)} per_hh nodes)")

    for ho in ho_list:
        # Scope assignment for root node
        existing_scope = (
            session.query(NodeHouseholdScope)
            .filter(
                NodeHouseholdScope.workflow_node_id == snap_node_id,
                NodeHouseholdScope.ho_id == ho.id,
            )
            .first()
        )
        if existing_scope is None:
            scope = NodeHouseholdScope(
                id=uuid.uuid4(),
                workflow_node_id=snap_node_id,
                ho_id=ho.id,
            )
            session.add(scope)

        # Task instances for all per_household nodes
        for node in per_hh_nodes:
            existing_task = (
                session.query(TaskInstance)
                .filter(
                    TaskInstance.workflow_node_id == node.id,
                    TaskInstance.ho_id == ho.id,
                )
                .first()
            )
            if existing_task is None:
                task = TaskInstance(
                    id=uuid.uuid4(),
                    ho_so_id=ho_so_id,
                    workflow_node_id=node.id,
                    ho_id=ho.id,
                    status=TaskStatusEnum.dang_thuc_hien,
                )
                session.add(task)

    session.flush()


def complete_tasks_for_codes(
    session: Session,
    ho_so_id: uuid.UUID,
    ho_id: uuid.UUID,
    codes: List[str],
) -> None:
    """Mark specified task codes as hoan_thanh for a given ho."""
    for code in codes:
        snap_node = (
            session.query(HoSoWorkflowNode)
            .filter(
                HoSoWorkflowNode.ho_so_id == ho_so_id,
                HoSoWorkflowNode.code == code,
            )
            .first()
        )
        if snap_node is None:
            print(f"  WARNING: snap node for code {code} not found")
            continue

        # Find task for this ho or None (for non-per_household tasks)
        task = (
            session.query(TaskInstance)
            .filter(
                TaskInstance.workflow_node_id == snap_node.id,
                TaskInstance.ho_id == (ho_id if snap_node.per_household else None),
            )
            .first()
        )
        if task:
            task.status = TaskStatusEnum.hoan_thanh
            task.completed_at = datetime.utcnow()
        else:
            print(f"  WARNING: task for code {code} not found")

    # Also mark non-per_hh C001-C007 as done (they don't have ho_id)
    session.flush()


# ── Main seed logic ───────────────────────────────────────────────────────────

def seed():
    print("=" * 60)
    print("OPC GPMB Seed Script")
    print("=" * 60)

    Base.metadata.create_all(engine)

    with SessionLocal() as session:
        # ── 1. Workflow Template ─────────────────────────────────────────

        print("\n[1] Creating WorkflowTemplate...")
        existing_template = (
            session.query(WorkflowTemplate)
            .filter(WorkflowTemplate.name == "Quy trình GPMB chuẩn")
            .first()
        )
        if existing_template:
            print("  Template already exists, skipping.")
            template = existing_template
        else:
            template = WorkflowTemplate(
                id=uuid.uuid4(),
                name="Quy trình GPMB chuẩn",
                is_active=True,
            )
            session.add(template)
            session.flush()
            print(f"  Created template: {template.name} ({template.id})")

            # ── Read Excel Sheet 1 ──────────────────────────────────────
            print("\n[2] Reading workflow nodes from Excel...")
            wb = openpyxl.load_workbook(EXCEL_PATH)
            ws1 = wb.worksheets[0]

            # Node structure based on actual Excel analysis:
            # Row 4: "I" → GD2 (level 1, per_household=False)
            #   Rows 6-12: C001-C007 (level 2, per_household=False)
            #   Sub-group: "Kiểm đếm & Thông báo từng hộ" (level 2, per_household=True)
            #     Rows 13-71: C008-C063 (level 3, per_household=True)
            # Row 72: "II" → GD3 (level 1, per_household=True)
            #   Rows 73-86: C064-C077 (level 2, per_household=True)
            # Row 88: "III" → GD4 (level 1, per_household=True)
            #   Rows 89-121: C078-C110 (level 2, per_household=True)

            # Parse all rows
            all_rows = []
            for i, row in enumerate(ws1.iter_rows(min_row=4, max_row=130, values_only=True), start=4):
                col_a, col_b, col_c, col_d, col_e, col_f = (
                    row[0], row[1], row[2], row[3], row[4], row[5]
                )
                col_k = row[10]  # planned_days
                if col_a is not None or col_b is not None:
                    all_rows.append({
                        "row": i,
                        "a": col_a,
                        "b": col_b,
                        "c": col_c,
                        "d": col_d,
                        "e": col_e,
                        "f": col_f,
                        "k": col_k,
                    })

            order_counter = 0

            # ── GD2 (Giai đoạn 2) ───────────────────────────────────────
            gd2_name = "GIAI ĐOẠN 2: THÔNG BÁO THU HỒI ĐẤT + KIỂM ĐẾM, XÁC NHẬN THÔNG TIN"
            gd2 = create_wf_node(
                session,
                template.id,
                name=gd2_name,
                level=1,
                order=order_counter,
                per_household=False,
            )
            order_counter += 1
            print(f"  Created GD2: {gd2.name[:50]}")

            # C001–C007: level 2, non-per_household, children of GD2
            c001_c007_codes = ["C001", "C002", "C003", "C004", "C005", "C006", "C007"]
            for r in all_rows:
                if r["b"] in c001_c007_codes:
                    planned = int(r["k"]) if r["k"] and isinstance(r["k"], (int, float)) else None
                    create_wf_node(
                        session,
                        template.id,
                        name=str(r["c"]).strip(),
                        level=2,
                        order=order_counter,
                        code=r["b"],
                        parent_id=gd2.id,
                        per_household=False,
                        planned_days=planned,
                        legal_basis=str(r["d"]).strip() if r["d"] else None,
                        org_in_charge=str(r["e"]).strip() if r["e"] else None,
                        org_coordinate=str(r["f"]).strip() if r["f"] else None,
                    )
                    order_counter += 1

            # Sub-group for per_household nodes under GD2
            kiem_dem_group = create_wf_node(
                session,
                template.id,
                name="Kiểm đếm & Thông báo từng hộ",
                level=2,
                order=order_counter,
                parent_id=gd2.id,
                per_household=True,
            )
            order_counter += 1
            print(f"  Created sub-group: {kiem_dem_group.name}")

            # C008-C063: level 3, per_household=True, parent=kiem_dem_group
            gd2_per_hh_codes = [f"C{str(i).zfill(3)}" for i in range(8, 64)]
            for r in all_rows:
                if r["b"] in gd2_per_hh_codes:
                    planned = None
                    if r["k"] and isinstance(r["k"], (int, float)):
                        planned = int(r["k"])
                    create_wf_node(
                        session,
                        template.id,
                        name=str(r["c"]).strip(),
                        level=3,
                        order=order_counter,
                        code=r["b"],
                        parent_id=kiem_dem_group.id,
                        per_household=True,
                        planned_days=planned,
                        legal_basis=str(r["d"]).strip() if r["d"] else None,
                        org_in_charge=str(r["e"]).strip() if r["e"] else None,
                        org_coordinate=str(r["f"]).strip() if r["f"] else None,
                    )
                    order_counter += 1

            # ── GD3 (Giai đoạn 3) ───────────────────────────────────────
            gd3_name = "GIAI ĐOẠN 3: LẬP, CÔNG KHAI, TRÌNH THẨM ĐỊNH, PHÊ DUYỆT PHƯƠNG ÁN BỒI THƯỜNG"
            gd3 = create_wf_node(
                session,
                template.id,
                name=gd3_name,
                level=1,
                order=order_counter,
                per_household=True,
            )
            order_counter += 1
            print(f"  Created GD3: {gd3.name[:50]}")

            # C064-C077: level 2, per_household=True, parent=GD3
            gd3_codes = [f"C{str(i).zfill(3)}" for i in range(64, 78)]
            for r in all_rows:
                if r["b"] in gd3_codes:
                    planned = None
                    if r["k"] and isinstance(r["k"], (int, float)):
                        planned = int(r["k"])
                    create_wf_node(
                        session,
                        template.id,
                        name=str(r["c"]).strip(),
                        level=2,
                        order=order_counter,
                        code=r["b"],
                        parent_id=gd3.id,
                        per_household=True,
                        planned_days=planned,
                        legal_basis=str(r["d"]).strip() if r["d"] else None,
                        org_in_charge=str(r["e"]).strip() if r["e"] else None,
                        org_coordinate=str(r["f"]).strip() if r["f"] else None,
                    )
                    order_counter += 1

            # ── GD4 (Giai đoạn 4) ───────────────────────────────────────
            gd4_name = "GIAI ĐOẠN 4: CHI TRẢ VÀ BÀN GIAO MẶT BẰNG"
            gd4 = create_wf_node(
                session,
                template.id,
                name=gd4_name,
                level=1,
                order=order_counter,
                per_household=True,
            )
            order_counter += 1
            print(f"  Created GD4: {gd4.name[:50]}")

            # C078-C110: level 2, per_household=True, parent=GD4
            gd4_codes = [f"C{str(i).zfill(3)}" for i in range(78, 111)]
            for r in all_rows:
                if r["b"] in gd4_codes:
                    planned = None
                    if r["k"] and isinstance(r["k"], (int, float)):
                        planned = int(r["k"])
                    create_wf_node(
                        session,
                        template.id,
                        name=str(r["c"]).strip(),
                        level=2,
                        order=order_counter,
                        code=r["b"],
                        parent_id=gd4.id,
                        per_household=True,
                        planned_days=planned,
                        legal_basis=str(r["d"]).strip() if r["d"] else None,
                        org_in_charge=str(r["e"]).strip() if r["e"] else None,
                        org_coordinate=str(r["f"]).strip() if r["f"] else None,
                    )
                    order_counter += 1

            session.commit()
            print(f"  Workflow nodes committed. Total order counter: {order_counter}")

        # ── 2. Users ────────────────────────────────────────────────────

        print("\n[3] Creating users...")
        users_data = [
            {
                "username": "admin",
                "password": "Admin@123",
                "full_name": "Admin Hệ thống",
                "role": RoleEnum.admin,
            },
            {
                "username": "cbcq",
                "password": "Cbcq@123",
                "full_name": "Nguyễn Văn Khánh",
                "role": RoleEnum.cbcq,
            },
            {
                "username": "ketoan",
                "password": "Ketoan@123",
                "full_name": "Trần Thị Hoa",
                "role": RoleEnum.ke_toan,
            },
            {
                "username": "gddh",
                "password": "Gddh@123",
                "full_name": "Lê Văn Minh",
                "role": RoleEnum.gddh,
            },
        ]

        user_map: Dict[str, User] = {}
        for u_data in users_data:
            existing_user = (
                session.query(User)
                .filter(User.username == u_data["username"])
                .first()
            )
            if existing_user:
                print(f"  User '{u_data['username']}' already exists, skipping.")
                user_map[u_data["username"]] = existing_user
            else:
                user = User(
                    id=uuid.uuid4(),
                    username=u_data["username"],
                    password_hash=pwd_context.hash(u_data["password"]),
                    full_name=u_data["full_name"],
                    role=u_data["role"],
                    active=True,
                )
                session.add(user)
                user_map[u_data["username"]] = user
                print(f"  Created user: {u_data['username']} ({u_data['role'].value})")

        session.flush()

        # ── 3. HoSoGPMB ─────────────────────────────────────────────────

        print("\n[4] Creating HoSoGPMB...")
        existing_hs = (
            session.query(HoSoGPMB).filter(HoSoGPMB.code == "HS-202504-001").first()
        )
        if existing_hs:
            print("  HoSo already exists, skipping.")
            ho_so = existing_hs
        else:
            cbcq_user = user_map["cbcq"]
            ho_so = HoSoGPMB(
                id=uuid.uuid4(),
                code="HS-202504-001",
                name="CCN Hữu Bằng",
                dia_chi="Thôn Hữu Bằng, Thạch Thất, Hà Nội",
                status=HoSoStatusEnum.thuc_hien,
                cbcq_id=cbcq_user.id,
                template_id=template.id,
            )
            session.add(ho_so)
            session.flush()
            print(f"  Created HoSo: {ho_so.name} ({ho_so.id})")

            # Snapshot workflow
            print("  Snapshotting workflow...")
            id_map = snapshot_workflow(session, ho_so.id, template.id)
            print(f"  Snapshotted {len(id_map)} nodes.")

            # Generate non-per_household tasks
            task_count = generate_non_per_hh_tasks(session, ho_so.id)
            print(f"  Generated {task_count} non-per-household tasks.")

            session.commit()

        # ── 4. Ho (households) from Sheet 2 ────────────────────────────

        print("\n[5] Importing households from Sheet 2...")
        existing_ho_count = (
            session.query(Ho).filter(Ho.ho_so_id == ho_so.id).count()
        )
        if existing_ho_count > 0:
            print(f"  Households already exist ({existing_ho_count}), skipping import.")
        else:
            wb = openpyxl.load_workbook(EXCEL_PATH)
            ws2 = wb.worksheets[1]
            ho_list_objects = []
            inserted = 0
            for row_idx, row in enumerate(
                ws2.iter_rows(min_row=5, max_row=457, values_only=True), start=5
            ):
                stt, ma_ho, loai_dat, ten_chu_ho, dia_chi, thua, dien_tich = (
                    row[0],
                    row[1],
                    row[2],
                    row[3],
                    row[4],
                    row[5],
                    row[6],
                )
                if stt is None or ma_ho is None:
                    continue

                dien_tich_float = None
                if dien_tich is not None:
                    try:
                        dien_tich_float = float(dien_tich)
                    except (ValueError, TypeError):
                        pass

                ho = Ho(
                    id=uuid.uuid4(),
                    ho_so_id=ho_so.id,
                    ma_ho=str(ma_ho).strip(),
                    ten_chu_ho=str(ten_chu_ho).strip() if ten_chu_ho else "Không rõ",
                    dia_chi=str(dia_chi).strip() if dia_chi else None,
                    loai_dat=str(loai_dat).strip() if loai_dat else None,
                    thua=str(thua).strip() if thua else None,
                    dien_tich=dien_tich_float,
                    status=HoStatusEnum.moi,
                )
                session.add(ho)
                ho_list_objects.append(ho)
                inserted += 1

                if inserted % 100 == 0:
                    session.flush()
                    print(f"  ... {inserted} households flushed")

            session.flush()
            session.commit()
            print(f"  Inserted {inserted} households.")

        # ── 5. Scope assignment for all per_household nodes ─────────────

        print("\n[6] Assigning households to per_household nodes...")

        # Get all ho for this ho_so
        ho_list: List[Ho] = (
            session.query(Ho).filter(Ho.ho_so_id == ho_so.id).all()
        )
        print(f"  Total households: {len(ho_list)}")

        # Check if scope already assigned
        existing_scope_count = (
            session.query(NodeHouseholdScope)
            .join(HoSoWorkflowNode, NodeHouseholdScope.workflow_node_id == HoSoWorkflowNode.id)
            .filter(HoSoWorkflowNode.ho_so_id == ho_so.id)
            .count()
        )
        if existing_scope_count > 0:
            print(f"  Scope already assigned ({existing_scope_count} records), skipping.")
        else:
            # Find the top-level per_household nodes in the snapshot
            # These are: kiem_dem_group snap, GD3 snap, GD4 snap
            snap_per_hh_roots = (
                session.query(HoSoWorkflowNode)
                .filter(
                    HoSoWorkflowNode.ho_so_id == ho_so.id,
                    HoSoWorkflowNode.per_household == True,
                    HoSoWorkflowNode.level.in_([1, 2]),  # root or sub-group
                )
                .all()
            )

            # We only want the 3 top-level per_hh entry points:
            # - kiem_dem_group (level 2, parent=GD2 snap)
            # - GD3 snap (level 1)
            # - GD4 snap (level 1)
            # To find them: level 1 nodes with per_household=True, plus
            # level 2 nodes with per_household=True whose parent is per_household=False
            entry_nodes = []
            for node in snap_per_hh_roots:
                if node.level == 1:
                    entry_nodes.append(node)
                elif node.level == 2:
                    parent = session.get(HoSoWorkflowNode, node.parent_id)
                    if parent and not parent.per_household:
                        entry_nodes.append(node)

            print(f"  Entry per_hh nodes: {[n.name[:40] for n in entry_nodes]}")

            for entry_node in entry_nodes:
                assign_all_households_to_node(session, ho_so.id, entry_node.id, ho_list)
                session.flush()
                print(f"  Assigned all households to: {entry_node.name[:50]}")

            session.commit()
            print("  Scope assignment complete.")

        # ── 6. Demo data: complete C001-C010 for HB001 ─────────────────

        print("\n[7] Setting demo completed tasks for HB001...")
        hb001 = (
            session.query(Ho)
            .filter(Ho.ho_so_id == ho_so.id, Ho.ma_ho == "HB001")
            .first()
        )
        if hb001 is None:
            print("  HB001 not found, skipping demo data.")
        else:
            # C001-C007 are non-per_household, C008-C010 are per_household
            non_per_hh_codes = ["C001", "C002", "C003", "C004", "C005", "C006", "C007"]
            per_hh_codes = ["C008", "C009", "C010"]

            # Mark non-per_household tasks as done
            for code in non_per_hh_codes:
                snap_node = (
                    session.query(HoSoWorkflowNode)
                    .filter(
                        HoSoWorkflowNode.ho_so_id == ho_so.id,
                        HoSoWorkflowNode.code == code,
                    )
                    .first()
                )
                if snap_node:
                    task = (
                        session.query(TaskInstance)
                        .filter(
                            TaskInstance.workflow_node_id == snap_node.id,
                            TaskInstance.ho_id == None,
                        )
                        .first()
                    )
                    if task and task.status != TaskStatusEnum.hoan_thanh:
                        task.status = TaskStatusEnum.hoan_thanh
                        task.completed_at = datetime.utcnow()
                        print(f"  Completed {code} (non-per_hh)")

            # Mark per_household tasks as done for HB001
            for code in per_hh_codes:
                snap_node = (
                    session.query(HoSoWorkflowNode)
                    .filter(
                        HoSoWorkflowNode.ho_so_id == ho_so.id,
                        HoSoWorkflowNode.code == code,
                    )
                    .first()
                )
                if snap_node:
                    task = (
                        session.query(TaskInstance)
                        .filter(
                            TaskInstance.workflow_node_id == snap_node.id,
                            TaskInstance.ho_id == hb001.id,
                        )
                        .first()
                    )
                    if task and task.status != TaskStatusEnum.hoan_thanh:
                        task.status = TaskStatusEnum.hoan_thanh
                        task.completed_at = datetime.utcnow()
                        print(f"  Completed {code} for HB001")

            # Update HB001 status to dang_xu_ly
            if hb001.status == HoStatusEnum.moi:
                hb001.status = HoStatusEnum.dang_xu_ly

            session.commit()
            print("  Demo data set.")

        print("\n" + "=" * 60)
        print("Seed complete!")
        print("=" * 60)
        print("\nUser credentials:")
        print("  admin / Admin@123")
        print("  cbcq / Cbcq@123")
        print("  ketoan / Ketoan@123")
        print("  gddh / Gddh@123")
        print("\nHo So: HS-202504-001 (CCN Hữu Bằng)")
        print(f"Total households: {session.query(Ho).filter(Ho.ho_so_id == ho_so.id).count()}")


if __name__ == "__main__":
    seed()
