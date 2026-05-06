from __future__ import annotations

import io
import uuid
from typing import Optional, List, Dict, Any

from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, delete
from sqlalchemy.orm import selectinload
from pydantic import BaseModel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from ...db.session import get_db
from ...db.models import Ho, HoDatInfo, HoSoGPMB, HoStatusEnum, TaskInstance, HoSoChiTra, User, RoleEnum
from ..deps import get_current_user, require_roles


class HoStatusUpdate(BaseModel):
    status: str

router = APIRouter()


# ── Schemas ──────────────────────────────────────────────────────────────────

# Vietnamese land type catalog (Luật Đất đai 2013 / 2024)
LOAI_DAT_CATALOG: List[Dict[str, str]] = [
    # Nhóm đất nông nghiệp
    {"code": "LUC", "name": "Đất trồng lúa nước"},
    {"code": "LNC", "name": "Đất trồng lúa nương"},
    {"code": "BHK", "name": "Đất bằng trồng cây hàng năm khác"},
    {"code": "NHK", "name": "Đất nương rẫy trồng cây hàng năm khác"},
    {"code": "CLN", "name": "Đất trồng cây lâu năm"},
    {"code": "RSX", "name": "Đất rừng sản xuất"},
    {"code": "RPH", "name": "Đất rừng phòng hộ"},
    {"code": "RDD", "name": "Đất rừng đặc dụng"},
    {"code": "NTS", "name": "Đất nuôi trồng thủy sản"},
    {"code": "LMU", "name": "Đất làm muối"},
    {"code": "NKH", "name": "Đất nông nghiệp khác"},
    # Nhóm đất phi nông nghiệp
    {"code": "ONT", "name": "Đất ở tại nông thôn"},
    {"code": "ODT", "name": "Đất ở tại đô thị"},
    {"code": "TSC", "name": "Đất xây dựng trụ sở cơ quan"},
    {"code": "DTS", "name": "Đất xây dựng trụ sở của tổ chức sự nghiệp"},
    {"code": "SKC", "name": "Đất cơ sở sản xuất phi nông nghiệp"},
    {"code": "SKS", "name": "Đất sử dụng cho hoạt động khoáng sản"},
    {"code": "CSD", "name": "Đất sử dụng vào mục đích công cộng"},
    {"code": "TIN", "name": "Đất tín ngưỡng"},
    {"code": "TON", "name": "Đất tôn giáo"},
    {"code": "NTD", "name": "Đất nghĩa trang, nghĩa địa"},
    {"code": "MNC", "name": "Đất có mặt nước chuyên dùng"},
    {"code": "PNK", "name": "Đất phi nông nghiệp khác"},
    # Nhóm đất chưa sử dụng
    {"code": "DCS", "name": "Đất chưa sử dụng"},
]

VALID_LOAI_DAT_CODES = {item["code"] for item in LOAI_DAT_CATALOG}


class HoDatInfoCreate(BaseModel):
    loai_dat: str
    so_tien: Optional[float] = None
    ghi_chu: Optional[str] = None


class HoCreate(BaseModel):
    ma_ho: str
    ten_chu_ho: str
    loai_doi_tuong: Optional[str] = None  # ca_nhan | to_chuc
    dia_chi: Optional[str] = None
    so_dien_thoai: Optional[str] = None
    thua: Optional[str] = None  # Số thửa
    so_to_ban_do: Optional[str] = None
    dien_tich: Optional[float] = None
    ty_le_thu_hoi: Optional[float] = None
    cccd: Optional[str] = None
    dkkd_mst: Optional[str] = None
    ghi_chu: Optional[str] = None
    dat_info: Optional[List[HoDatInfoCreate]] = None


class HoUpdate(BaseModel):
    ma_ho: Optional[str] = None
    ten_chu_ho: Optional[str] = None
    loai_doi_tuong: Optional[str] = None
    dia_chi: Optional[str] = None
    so_dien_thoai: Optional[str] = None
    thua: Optional[str] = None
    so_to_ban_do: Optional[str] = None
    dien_tich: Optional[float] = None
    ty_le_thu_hoi: Optional[float] = None
    cccd: Optional[str] = None
    dkkd_mst: Optional[str] = None
    ghi_chu: Optional[str] = None
    dat_info: Optional[List[HoDatInfoCreate]] = None  # replaces all existing dat_info if provided


def _dat_info_to_dict(d: HoDatInfo) -> Dict[str, Any]:
    return {
        "id": str(d.id),
        "loai_dat": d.loai_dat,
        "so_tien": d.so_tien,
        "ghi_chu": d.ghi_chu,
    }


def ho_to_dict(ho: Ho) -> Dict[str, Any]:
    return {
        "id": str(ho.id),
        "ho_so_id": str(ho.ho_so_id),
        "ma_ho": ho.ma_ho,
        "ten_chu_ho": ho.ten_chu_ho,
        "loai_doi_tuong": ho.loai_doi_tuong,
        "dia_chi": ho.dia_chi,
        "so_dien_thoai": ho.so_dien_thoai,
        "thua": ho.thua,
        "so_to_ban_do": ho.so_to_ban_do,
        "dien_tich": ho.dien_tich,
        "ty_le_thu_hoi": ho.ty_le_thu_hoi,
        "cccd": ho.cccd,
        "dkkd_mst": ho.dkkd_mst,
        "ghi_chu": ho.ghi_chu,
        "dat_info": [_dat_info_to_dict(d) for d in (ho.dat_info or [])],
        # legacy field kept for compatibility (may be populated from old imports)
        "loai_dat": ho.loai_dat,
        "status": ho.status.value,
        "created_at": ho.created_at.isoformat(),
        "updated_at": ho.updated_at.isoformat(),
    }


async def _get_ho_so_or_404(ho_so_id: str, db: AsyncSession) -> HoSoGPMB:
    result = await db.execute(
        select(HoSoGPMB).where(HoSoGPMB.id == uuid.UUID(ho_so_id))
    )
    ho_so = result.scalar_one_or_none()
    if ho_so is None:
        raise HTTPException(status_code=404, detail="Ho so not found")
    return ho_so


# ── Endpoints ────────────────────────────────────────────────────────────────

@router.get("/{ho_so_id}/ho/{ho_id}")
async def get_ho(
    ho_so_id: str,
    ho_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await _get_ho_so_or_404(ho_so_id, db)
    result = await db.execute(
        select(Ho).options(selectinload(Ho.dat_info)).where(
            Ho.id == uuid.UUID(ho_id),
            Ho.ho_so_id == uuid.UUID(ho_so_id),
        )
    )
    ho = result.scalar_one_or_none()
    if ho is None:
        raise HTTPException(status_code=404, detail="Hộ không tồn tại")
    return ho_to_dict(ho)


@router.get("/{ho_so_id}/ho")
async def list_ho(
    ho_so_id: str,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=500),
    status: Optional[str] = None,
    search: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await _get_ho_so_or_404(ho_so_id, db)

    q = select(Ho).options(selectinload(Ho.dat_info)).where(Ho.ho_so_id == uuid.UUID(ho_so_id))
    if status:
        try:
            status_enum = HoStatusEnum(status)
            q = q.where(Ho.status == status_enum)
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Invalid status: {status}")
    if search:
        q = q.where(
            Ho.ten_chu_ho.ilike(f"%{search}%") | Ho.ma_ho.ilike(f"%{search}%")
        )

    count_result = await db.execute(select(func.count()).select_from(
        select(Ho).where(Ho.ho_so_id == uuid.UUID(ho_so_id)).subquery()
    ))
    total = count_result.scalar()

    q = q.order_by(Ho.ma_ho).offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(q)
    items = result.scalars().all()

    return {
        "items": [ho_to_dict(h) for h in items],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@router.post("/{ho_so_id}/ho", status_code=status.HTTP_201_CREATED)
async def create_ho(
    ho_so_id: str,
    body: HoCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles(RoleEnum.admin, RoleEnum.cbcq)),
):
    await _get_ho_so_or_404(ho_so_id, db)

    # Check ma_ho uniqueness within ho_so
    existing = await db.execute(
        select(Ho).where(
            Ho.ho_so_id == uuid.UUID(ho_so_id),
            Ho.ma_ho == body.ma_ho,
        )
    )
    if existing.scalar_one_or_none():
        raise HTTPException(
            status_code=400, detail=f"Ma ho '{body.ma_ho}' already exists"
        )

    ho = Ho(
        ho_so_id=uuid.UUID(ho_so_id),
        ma_ho=body.ma_ho,
        ten_chu_ho=body.ten_chu_ho,
        loai_doi_tuong=body.loai_doi_tuong,
        dia_chi=body.dia_chi,
        so_dien_thoai=body.so_dien_thoai,
        thua=body.thua,
        so_to_ban_do=body.so_to_ban_do,
        dien_tich=body.dien_tich,
        ty_le_thu_hoi=body.ty_le_thu_hoi,
        cccd=body.cccd,
        dkkd_mst=body.dkkd_mst,
        ghi_chu=body.ghi_chu,
        status=HoStatusEnum.moi,
    )
    db.add(ho)
    await db.flush()  # get ho.id before adding children

    for dat in (body.dat_info or []):
        db.add(HoDatInfo(ho_id=ho.id, loai_dat=dat.loai_dat, so_tien=dat.so_tien, ghi_chu=dat.ghi_chu))

    await db.commit()
    await db.refresh(ho)
    # reload dat_info
    result2 = await db.execute(select(Ho).options(selectinload(Ho.dat_info)).where(Ho.id == ho.id))
    ho = result2.scalar_one()
    return ho_to_dict(ho)


@router.get("/{ho_so_id}/ho/import/template")
async def download_import_template(
    ho_so_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Download an xlsx template for importing households."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Import Ho"

    headers = ["ma_ho", "ten_chu_ho", "dia_chi", "loai_dat", "thua", "dien_tich"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Example row
    ws.append(["HB001", "Nguyễn Văn A", "Thôn Hữu Bằng", "LUC", "123", 500.0])

    # Adjust column widths
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=import_ho_template.xlsx"},
    )


@router.post("/{ho_so_id}/ho/import")
async def import_ho(
    ho_so_id: str,
    file: UploadFile = File(...),
    confirm: bool = Query(False),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles(RoleEnum.admin, RoleEnum.cbcq)),
):
    """
    Parse xlsx file and validate rows.
    If confirm=false, return preview with valid/error rows.
    If confirm=true, insert valid rows into DB.
    """
    await _get_ho_so_or_404(ho_so_id, db)

    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Cannot read xlsx file: {e}")

    ws = wb.active
    valid_rows: List[Dict] = []
    error_rows: List[Dict] = []

    # Get existing ma_ho set for duplicate detection
    existing_result = await db.execute(
        select(Ho.ma_ho).where(Ho.ho_so_id == uuid.UUID(ho_so_id))
    )
    existing_ma_ho = set(row[0] for row in existing_result.all())
    seen_in_file: set = set()

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue

        ma_ho = str(row[0]).strip() if row[0] is not None else ""
        ten_chu_ho = str(row[1]).strip() if row[1] is not None else ""
        dia_chi = str(row[2]).strip() if row[2] is not None else None
        loai_dat = str(row[3]).strip() if row[3] is not None else None
        thua = str(row[4]).strip() if row[4] is not None else None
        dien_tich = None
        if row[5] is not None:
            try:
                dien_tich = float(row[5])
            except (ValueError, TypeError):
                pass

        errors = []
        if not ma_ho:
            errors.append("ma_ho is required")
        if not ten_chu_ho:
            errors.append("ten_chu_ho is required")
        if ma_ho in existing_ma_ho:
            errors.append(f"ma_ho '{ma_ho}' already exists in database")
        if ma_ho in seen_in_file:
            errors.append(f"ma_ho '{ma_ho}' is duplicated in file")

        row_data = {
            "row": row_idx,
            "ma_ho": ma_ho,
            "ten_chu_ho": ten_chu_ho,
            "dia_chi": dia_chi,
            "loai_dat": loai_dat,
            "thua": thua,
            "dien_tich": dien_tich,
        }

        if errors:
            error_rows.append({**row_data, "errors": errors})
        else:
            valid_rows.append(row_data)
            seen_in_file.add(ma_ho)

    imported = False
    if confirm and valid_rows:
        for row_data in valid_rows:
            ho = Ho(
                ho_so_id=uuid.UUID(ho_so_id),
                ma_ho=row_data["ma_ho"],
                ten_chu_ho=row_data["ten_chu_ho"],
                dia_chi=row_data["dia_chi"],
                loai_dat=row_data["loai_dat"],
                thua=row_data["thua"],
                dien_tich=row_data["dien_tich"],
                status=HoStatusEnum.moi,
            )
            db.add(ho)
        await db.commit()
        imported = True

    return {
        "valid": valid_rows,
        "errors": error_rows,
        "imported": imported,
        "valid_count": len(valid_rows),
        "error_count": len(error_rows),
    }


@router.patch("/{ho_so_id}/ho/{ho_id}/status")
async def update_ho_status(
    ho_so_id: str,
    ho_id: str,
    body: HoStatusUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles(RoleEnum.admin, RoleEnum.cbcq)),
):
    """Manually transition a ho's status. Only 'da_thong_nhat' is allowed via this endpoint."""
    await _get_ho_so_or_404(ho_so_id, db)
    ho_result = await db.execute(
        select(Ho).where(
            Ho.id == uuid.UUID(ho_id),
            Ho.ho_so_id == uuid.UUID(ho_so_id),
        )
    )
    ho = ho_result.scalar_one_or_none()
    if ho is None:
        raise HTTPException(status_code=404, detail="Ho not found")

    try:
        new_status = HoStatusEnum(body.status)
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Invalid status: {body.status}")

    allowed_manual = {HoStatusEnum.da_thong_nhat}
    if new_status not in allowed_manual:
        raise HTTPException(
            status_code=400,
            detail=f"Status '{new_status.value}' cannot be set manually",
        )

    valid_transitions = {
        HoStatusEnum.dang_xu_ly: [HoStatusEnum.da_thong_nhat],
    }
    if new_status not in valid_transitions.get(ho.status, []):
        raise HTTPException(
            status_code=400,
            detail=f"Cannot transition from {ho.status.value} to {new_status.value}",
        )

    ho.status = new_status
    await db.commit()
    await db.refresh(ho)
    return ho_to_dict(ho)


# S2-BE-03 — PATCH /ho-so/{ho_so_id}/ho/{ho_id}
@router.patch("/{ho_so_id}/ho/{ho_id}")
async def update_ho(
    ho_so_id: str,
    ho_id: str,
    body: HoUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles(RoleEnum.admin, RoleEnum.cbcq)),
):
    await _get_ho_so_or_404(ho_so_id, db)

    ho_result = await db.execute(
        select(Ho).where(
            Ho.id == uuid.UUID(ho_id),
            Ho.ho_so_id == uuid.UUID(ho_so_id),
        )
    )
    ho = ho_result.scalar_one_or_none()
    if ho is None:
        raise HTTPException(status_code=404, detail="Hộ không tồn tại")

    locked_statuses = {HoStatusEnum.da_thong_nhat, HoStatusEnum.da_chi_tra, HoStatusEnum.da_ban_giao}
    if ho.status in locked_statuses:
        raise HTTPException(
            status_code=409,
            detail=f"Không thể sửa hộ ở trạng thái '{ho.status.value}'"
        )

    update_data = body.model_dump(exclude_none=True)

    # Handle dat_info replacement separately
    new_dat_info = update_data.pop("dat_info", None)

    for field, value in update_data.items():
        setattr(ho, field, value)

    if new_dat_info is not None:
        # Delete all existing dat_info rows for this ho, then re-insert
        await db.execute(delete(HoDatInfo).where(HoDatInfo.ho_id == ho.id))
        for dat in new_dat_info:
            db.add(HoDatInfo(
                ho_id=ho.id,
                loai_dat=dat["loai_dat"],
                so_tien=dat.get("so_tien"),
                ghi_chu=dat.get("ghi_chu"),
            ))

    from datetime import datetime
    ho.updated_at = datetime.utcnow()
    await db.commit()
    # Reload with dat_info
    result2 = await db.execute(select(Ho).options(selectinload(Ho.dat_info)).where(Ho.id == ho.id))
    ho = result2.scalar_one()
    return ho_to_dict(ho)


# S2-BE-04 — DELETE /ho-so/{ho_so_id}/ho/{ho_id}
@router.delete("/{ho_so_id}/ho/{ho_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_ho(
    ho_so_id: str,
    ho_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles(RoleEnum.admin, RoleEnum.cbcq)),
):
    await _get_ho_so_or_404(ho_so_id, db)

    ho_result = await db.execute(
        select(Ho).where(
            Ho.id == uuid.UUID(ho_id),
            Ho.ho_so_id == uuid.UUID(ho_so_id),
        )
    )
    ho = ho_result.scalar_one_or_none()
    if ho is None:
        raise HTTPException(status_code=404, detail="Hộ không tồn tại")

    if ho.status != HoStatusEnum.moi:
        raise HTTPException(
            status_code=409, detail="Chỉ xoá được hộ chưa vào xử lý"
        )

    # Check task_instance records
    task_count = await db.scalar(
        select(func.count(TaskInstance.id)).where(TaskInstance.ho_id == ho.id)
    ) or 0
    if task_count > 0:
        raise HTTPException(
            status_code=409, detail="Hộ đã được gán vào quy trình"
        )

    # Check chi_tra records
    chi_tra_count = await db.scalar(
        select(func.count(HoSoChiTra.id)).where(HoSoChiTra.ho_id == ho.id)
    ) or 0
    if chi_tra_count > 0:
        raise HTTPException(
            status_code=409, detail="Hộ đã có chi trả, không thể xoá"
        )

    await db.delete(ho)
    await db.commit()
