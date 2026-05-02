from __future__ import annotations
import io
import os
import uuid
from datetime import date
from typing import Optional, List, Any, Dict
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Query, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import selectinload
from pydantic import BaseModel
import openpyxl
from openpyxl.styles import Font

from ...db.session import get_db
from ...db.models import WorkflowTemplate, WorkflowNode, WorkflowNodeDocument, User, RoleEnum
from ..deps import get_current_user, require_roles

UPLOAD_DIR = os.getenv("UPLOAD_DIR", "/tmp/odin-uploads")

router = APIRouter()


# ── Pydantic schemas ─────────────────────────────────────────────────────────

class WorkflowNodeCreate(BaseModel):
    template_id: str
    parent_id: Optional[str] = None
    code: Optional[str] = None
    name: str
    level: int
    order: int = 0
    planned_days: Optional[int] = None
    is_milestone: bool = False
    legal_basis: Optional[str] = None
    org_in_charge: Optional[str] = None
    org_coordinate: Optional[str] = None
    per_household: bool = False
    require_scan: bool = False
    field_so_vb: bool = False
    field_ngay_vb: bool = False
    field_loai_vb: bool = False
    field_gia_tri_trinh: bool = False
    field_gia_tri_duyet: bool = False
    field_ghi_chu: bool = False


class WorkflowNodeUpdate(BaseModel):
    code: Optional[str] = None
    name: Optional[str] = None
    level: Optional[int] = None
    order: Optional[int] = None
    planned_days: Optional[int] = None
    is_milestone: Optional[bool] = None
    legal_basis: Optional[str] = None
    org_in_charge: Optional[str] = None
    org_coordinate: Optional[str] = None
    per_household: Optional[bool] = None
    require_scan: Optional[bool] = None
    field_so_vb: Optional[bool] = None
    field_ngay_vb: Optional[bool] = None
    field_loai_vb: Optional[bool] = None
    field_gia_tri_trinh: Optional[bool] = None
    field_gia_tri_duyet: Optional[bool] = None
    field_ghi_chu: Optional[bool] = None


def node_to_dict(node: WorkflowNode) -> Dict[str, Any]:
    docs = []
    if hasattr(node, "documents") and node.documents:
        docs = [
            {
                "id": str(d.id),
                "ten_tai_lieu": d.ten_tai_lieu,
                "url": f"/uploads/{d.file_path}",
                "uploaded_at": d.uploaded_at.isoformat(),
            }
            for d in node.documents
        ]
    return {
        "id": str(node.id),
        "template_id": str(node.template_id),
        "parent_id": str(node.parent_id) if node.parent_id else None,
        "code": node.code,
        "name": node.name,
        "level": node.level,
        "order": node.order,
        "planned_days": node.planned_days,
        "is_milestone": node.is_milestone,
        "legal_basis": node.legal_basis,
        "org_in_charge": node.org_in_charge,
        "org_coordinate": node.org_coordinate,
        "per_household": node.per_household,
        "require_scan": node.require_scan,
        "field_so_vb": node.field_so_vb,
        "field_ngay_vb": node.field_ngay_vb,
        "field_loai_vb": node.field_loai_vb,
        "field_gia_tri_trinh": node.field_gia_tri_trinh,
        "field_gia_tri_duyet": node.field_gia_tri_duyet,
        "field_ghi_chu": node.field_ghi_chu,
        "documents": docs,
        "children": [],
    }


def build_tree(nodes: List[WorkflowNode]) -> List[Dict[str, Any]]:
    node_map: Dict[str, Dict] = {}
    roots: List[Dict] = []
    for node in sorted(nodes, key=lambda n: n.order):
        d = node_to_dict(node)
        node_map[str(node.id)] = d

    for node in sorted(nodes, key=lambda n: n.order):
        d = node_map[str(node.id)]
        if node.parent_id and str(node.parent_id) in node_map:
            node_map[str(node.parent_id)]["children"].append(d)
        else:
            roots.append(d)
    return roots


# ── Pydantic schemas (template) ───────────────────────────────────────────────

class WorkflowTemplateCreate(BaseModel):
    name: str


class WorkflowTemplateUpdate(BaseModel):
    name: Optional[str] = None
    is_active: Optional[bool] = None


# ── Template CRUD ─────────────────────────────────────────────────────────────

@router.get("/templates")
async def list_templates(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """List all workflow templates with node counts."""
    result = await db.execute(
        select(WorkflowTemplate).order_by(WorkflowTemplate.created_at.desc())
    )
    templates = result.scalars().all()

    out = []
    for t in templates:
        count_result = await db.execute(
            select(WorkflowNode).where(WorkflowNode.template_id == t.id)
        )
        node_count = len(count_result.scalars().all())
        out.append({
            "id": str(t.id),
            "name": t.name,
            "is_active": t.is_active,
            "created_at": t.created_at.isoformat(),
            "node_count": node_count,
        })
    return out


@router.post("/templates", status_code=201)
async def create_template(
    body: WorkflowTemplateCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles(RoleEnum.admin)),
):
    t = WorkflowTemplate(name=body.name, is_active=False)
    db.add(t)
    await db.commit()
    await db.refresh(t)
    return {"id": str(t.id), "name": t.name, "is_active": t.is_active, "created_at": t.created_at.isoformat(), "node_count": 0}


@router.get("/templates/{template_id}")
async def get_template(
    template_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    result = await db.execute(
        select(WorkflowTemplate).where(WorkflowTemplate.id == uuid.UUID(template_id))
    )
    template = result.scalar_one_or_none()
    if template is None:
        raise HTTPException(status_code=404, detail="Template not found")

    nodes_result = await db.execute(
        select(WorkflowNode)
        .where(WorkflowNode.template_id == template.id)
        .options(selectinload(WorkflowNode.documents))
    )
    nodes = list(nodes_result.scalars().all())

    return {
        "id": str(template.id),
        "name": template.name,
        "is_active": template.is_active,
        "created_at": template.created_at.isoformat(),
        "nodes": build_tree(nodes),
    }


@router.put("/templates/{template_id}")
async def update_template(
    template_id: str,
    body: WorkflowTemplateUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles(RoleEnum.admin)),
):
    result = await db.execute(
        select(WorkflowTemplate).where(WorkflowTemplate.id == uuid.UUID(template_id))
    )
    template = result.scalar_one_or_none()
    if template is None:
        raise HTTPException(status_code=404, detail="Template not found")

    if body.name is not None:
        template.name = body.name
    if body.is_active is not None:
        if body.is_active:
            # Deactivate all others first
            all_result = await db.execute(select(WorkflowTemplate))
            for t in all_result.scalars().all():
                t.is_active = False
        template.is_active = body.is_active

    await db.commit()
    await db.refresh(template)
    return {"id": str(template.id), "name": template.name, "is_active": template.is_active}


@router.delete("/templates/{template_id}", status_code=204)
async def delete_template(
    template_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles(RoleEnum.admin)),
):
    result = await db.execute(
        select(WorkflowTemplate).where(WorkflowTemplate.id == uuid.UUID(template_id))
    )
    template = result.scalar_one_or_none()
    if template is None:
        raise HTTPException(status_code=404, detail="Template not found")
    if template.is_active:
        raise HTTPException(status_code=400, detail="Không thể xóa quy trình đang hoạt động")
    await db.delete(template)
    await db.commit()


# ── Endpoints ────────────────────────────────────────────────────────────────

@router.get("/template")
async def get_active_template(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    result = await db.execute(
        select(WorkflowTemplate).where(WorkflowTemplate.is_active == True)  # noqa: E712
    )
    template = result.scalar_one_or_none()
    if template is None:
        raise HTTPException(status_code=404, detail="No active workflow template found")

    nodes_result = await db.execute(
        select(WorkflowNode)
        .where(WorkflowNode.template_id == template.id)
        .options(selectinload(WorkflowNode.documents))
    )
    nodes = list(nodes_result.scalars().all())

    return {
        "id": str(template.id),
        "name": template.name,
        "is_active": template.is_active,
        "created_at": template.created_at.isoformat(),
        "nodes": build_tree(nodes),
    }


@router.get("/nodes")
async def list_nodes(
    template_id: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    q = select(WorkflowNode).options(selectinload(WorkflowNode.documents))
    if template_id:
        q = q.where(WorkflowNode.template_id == uuid.UUID(template_id))
    result = await db.execute(q)
    nodes = list(result.scalars().all())
    return {"nodes": build_tree(nodes)}


@router.post("/nodes", status_code=status.HTTP_201_CREATED)
async def create_node(
    body: WorkflowNodeCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles(RoleEnum.admin)),
):
    node = WorkflowNode(
        template_id=uuid.UUID(body.template_id),
        parent_id=uuid.UUID(body.parent_id) if body.parent_id else None,
        code=body.code,
        name=body.name,
        level=body.level,
        order=body.order,
        planned_days=body.planned_days,
        is_milestone=body.is_milestone,
        legal_basis=body.legal_basis,
        org_in_charge=body.org_in_charge,
        org_coordinate=body.org_coordinate,
        per_household=body.per_household,
        require_scan=body.require_scan,
        field_so_vb=body.field_so_vb,
        field_ngay_vb=body.field_ngay_vb,
        field_loai_vb=body.field_loai_vb,
        field_gia_tri_trinh=body.field_gia_tri_trinh,
        field_gia_tri_duyet=body.field_gia_tri_duyet,
        field_ghi_chu=body.field_ghi_chu,
    )
    db.add(node)
    await db.commit()
    # Reload với selectinload để tránh MissingGreenlet khi node_to_dict truy cập documents
    result = await db.execute(
        select(WorkflowNode)
        .where(WorkflowNode.id == node.id)
        .options(selectinload(WorkflowNode.documents))
    )
    node = result.scalar_one()
    return node_to_dict(node)


@router.put("/nodes/{node_id}")
async def update_node(
    node_id: str,
    body: WorkflowNodeUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles(RoleEnum.admin)),
):
    result = await db.execute(
        select(WorkflowNode).where(WorkflowNode.id == uuid.UUID(node_id))
    )
    node = result.scalar_one_or_none()
    if node is None:
        raise HTTPException(status_code=404, detail="Node not found")

    for field, value in body.model_dump(exclude_none=True).items():
        setattr(node, field, value)

    await db.commit()
    # Reload với selectinload để tránh MissingGreenlet khi node_to_dict truy cập documents
    result = await db.execute(
        select(WorkflowNode)
        .where(WorkflowNode.id == node.id)
        .options(selectinload(WorkflowNode.documents))
    )
    node = result.scalar_one()
    return node_to_dict(node)


@router.delete("/nodes/{node_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_node(
    node_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles(RoleEnum.admin)),
):
    result = await db.execute(
        select(WorkflowNode).where(WorkflowNode.id == uuid.UUID(node_id))
    )
    node = result.scalar_one_or_none()
    if node is None:
        raise HTTPException(status_code=404, detail="Node not found")
    await db.delete(node)
    await db.commit()


# ── Node Documents ───────────────────────────────────────────────────────────

@router.get("/nodes/{node_id}/documents")
async def list_node_documents(
    node_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    result = await db.execute(
        select(WorkflowNode)
        .where(WorkflowNode.id == uuid.UUID(node_id))
        .options(selectinload(WorkflowNode.documents))
    )
    node = result.scalar_one_or_none()
    if node is None:
        raise HTTPException(status_code=404, detail="Node not found")
    return [
        {
            "id": str(d.id),
            "ten_tai_lieu": d.ten_tai_lieu,
            "url": f"/uploads/{d.file_path}",
            "uploaded_at": d.uploaded_at.isoformat(),
        }
        for d in node.documents
    ]


@router.post("/nodes/{node_id}/documents", status_code=status.HTTP_201_CREATED)
async def upload_node_document(
    node_id: str,
    ten_tai_lieu: str = Form(...),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles(RoleEnum.admin)),
):
    result = await db.execute(
        select(WorkflowNode).where(WorkflowNode.id == uuid.UUID(node_id))
    )
    node = result.scalar_one_or_none()
    if node is None:
        raise HTTPException(status_code=404, detail="Node not found")

    # Save file
    dest_dir = os.path.join(UPLOAD_DIR, "workflow-docs", node_id)
    os.makedirs(dest_dir, exist_ok=True)
    safe_name = f"{uuid.uuid4()}_{file.filename}"
    dest_path = os.path.join(dest_dir, safe_name)
    contents = await file.read()
    if len(contents) > 20 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="File quá lớn (tối đa 20MB)")
    with open(dest_path, "wb") as f:
        f.write(contents)

    relative_path = os.path.join("workflow-docs", node_id, safe_name)
    doc = WorkflowNodeDocument(
        node_id=uuid.UUID(node_id),
        ten_tai_lieu=ten_tai_lieu,
        file_path=relative_path,
        uploaded_by=current_user.id,
    )
    db.add(doc)
    await db.commit()
    await db.refresh(doc)
    return {
        "id": str(doc.id),
        "ten_tai_lieu": doc.ten_tai_lieu,
        "url": f"/uploads/{doc.file_path}",
        "uploaded_at": doc.uploaded_at.isoformat(),
    }


@router.delete("/nodes/{node_id}/documents/{doc_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_node_document(
    node_id: str,
    doc_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles(RoleEnum.admin)),
):
    result = await db.execute(
        select(WorkflowNodeDocument).where(
            WorkflowNodeDocument.id == uuid.UUID(doc_id),
            WorkflowNodeDocument.node_id == uuid.UUID(node_id),
        )
    )
    doc = result.scalar_one_or_none()
    if doc is None:
        raise HTTPException(status_code=404, detail="Document not found")
    # Remove file from disk (best effort)
    try:
        full_path = os.path.join(UPLOAD_DIR, doc.file_path)
        if os.path.exists(full_path):
            os.remove(full_path)
    except OSError:
        pass
    await db.delete(doc)
    await db.commit()


# ── Excel Import/Export constants ────────────────────────────────────────────

EXCEL_COLUMNS = [
    "code", "parent_code", "name", "planned_days", "per_household",
    "field_so_vb", "field_ngay_vb", "field_loai_vb",
    "field_gia_tri_trinh", "field_gia_tri_duyet", "field_ghi_chu",
    "require_scan", "legal_basis", "org_in_charge",
]

MAX_FILE_SIZE = 5 * 1024 * 1024  # 5 MB


def _bool_from_cell(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in ("true", "1", "yes", "x", "có")


# S2-BE-05 — POST /workflow/import-excel
@router.post("/import-excel")
async def import_workflow_excel(
    file: UploadFile = File(...),
    mode: str = Query("preview", regex="^(preview|confirm)$"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles(RoleEnum.admin)),
):
    # Validate filename
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="File phải là định dạng .xlsx")

    contents = await file.read()
    if len(contents) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="File vượt quá giới hạn 5MB")

    # Parse xlsx — data_only=True ensures no macro execution
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Không thể đọc file xlsx: {e}")

    ws = wb.active

    # Validate header row
    header_row = [str(cell.value).strip() if cell.value is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    if header_row != EXCEL_COLUMNS:
        raise HTTPException(
            status_code=400,
            detail=f"Header không đúng. Cần: {EXCEL_COLUMNS}, nhận được: {header_row}"
        )

    # Collect codes from DB for parent_code validation
    db_codes_result = await db.execute(select(WorkflowNode.code).where(WorkflowNode.code.isnot(None)))
    db_codes: set = set(row[0] for row in db_codes_result.all())

    rows_in_file: List[Dict] = []
    file_codes: set = set()
    preview_rows: List[Dict] = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(v for v in row if v is not None):
            continue

        code = str(row[0]).strip() if row[0] is not None else ""
        parent_code = str(row[1]).strip() if row[1] is not None else ""
        name = str(row[2]).strip() if row[2] is not None else ""

        errors = []
        if not code:
            errors.append("code là bắt buộc")

        if parent_code and parent_code not in file_codes and parent_code not in db_codes:
            errors.append(f"parent_code '{parent_code}' không tồn tại trong file hoặc DB")

        rows_in_file.append({
            "row_num": row_num,
            "code": code,
            "parent_code": parent_code or None,
            "name": name,
            "planned_days": int(row[3]) if row[3] is not None else None,
            "per_household": _bool_from_cell(row[4]),
            "field_so_vb": _bool_from_cell(row[5]),
            "field_ngay_vb": _bool_from_cell(row[6]),
            "field_loai_vb": _bool_from_cell(row[7]),
            "field_gia_tri_trinh": _bool_from_cell(row[8]),
            "field_gia_tri_duyet": _bool_from_cell(row[9]),
            "field_ghi_chu": _bool_from_cell(row[10]),
            "require_scan": _bool_from_cell(row[11]),
            "legal_basis": str(row[12]).strip() if row[12] is not None else None,
            "org_in_charge": str(row[13]).strip() if row[13] is not None else None,
            "errors": errors,
            "status": "error" if errors else "ok",
        })
        if not errors and code:
            file_codes.add(code)

    wb.close()

    valid_rows = [r for r in rows_in_file if r["status"] == "ok"]
    error_rows = [r for r in rows_in_file if r["status"] == "error"]

    summary = {
        "total": len(rows_in_file),
        "valid": len(valid_rows),
        "errors": len(error_rows),
    }

    if mode == "preview":
        return {
            "rows": [
                {
                    "stt": r["row_num"] - 1,  # 1-based data row index
                    "code": r["code"] or None,
                    "name": r["name"],
                    "parent_code": r.get("parent_code"),
                    "status": r["status"],
                    "detail": "; ".join(r["errors"]) if r["errors"] else None,
                }
                for r in rows_in_file
            ],
            "ok_count": summary["valid"],
            "error_count": summary["errors"],
        }

    # mode == "confirm": upsert
    # Get active template
    tmpl_result = await db.execute(
        select(WorkflowTemplate).where(WorkflowTemplate.is_active == True)  # noqa: E712
    )
    template = tmpl_result.scalar_one_or_none()
    if template is None:
        raise HTTPException(status_code=404, detail="Không tìm thấy workflow template đang hoạt động")

    inserted = 0
    updated = 0
    upsert_errors = 0

    # Build code->node map from DB for this template
    existing_nodes_result = await db.execute(
        select(WorkflowNode).where(
            WorkflowNode.template_id == template.id,
            WorkflowNode.code.isnot(None),
        )
    )
    existing_nodes: Dict[str, WorkflowNode] = {
        n.code: n for n in existing_nodes_result.scalars().all()
    }

    # Build code->id map (including newly inserted) for parent resolution
    code_to_id: Dict[str, uuid.UUID] = {code: n.id for code, n in existing_nodes.items()}
    # Also include DB nodes from other templates / no template for parent lookup
    all_db_codes_result = await db.execute(
        select(WorkflowNode.code, WorkflowNode.id).where(WorkflowNode.code.isnot(None))
    )
    global_code_map: Dict[str, uuid.UUID] = {row[0]: row[1] for row in all_db_codes_result.all()}

    new_node_map: Dict[str, uuid.UUID] = {}  # tracks nodes inserted in this run

    for row in valid_rows:
        try:
            # Resolve parent_id
            parent_id = None
            if row["parent_code"]:
                if row["parent_code"] in new_node_map:
                    parent_id = new_node_map[row["parent_code"]]
                elif row["parent_code"] in global_code_map:
                    parent_id = global_code_map[row["parent_code"]]

            if row["code"] in existing_nodes:
                # UPDATE
                node = existing_nodes[row["code"]]
                node.parent_id = parent_id
                node.name = row["name"]
                node.planned_days = row["planned_days"]
                node.per_household = row["per_household"]
                node.field_so_vb = row["field_so_vb"]
                node.field_ngay_vb = row["field_ngay_vb"]
                node.field_loai_vb = row["field_loai_vb"]
                node.field_gia_tri_trinh = row["field_gia_tri_trinh"]
                node.field_gia_tri_duyet = row["field_gia_tri_duyet"]
                node.field_ghi_chu = row["field_ghi_chu"]
                node.require_scan = row["require_scan"]
                node.legal_basis = row["legal_basis"]
                node.org_in_charge = row["org_in_charge"]
                updated += 1
                new_node_map[row["code"]] = node.id
            else:
                # INSERT — compute level based on parent
                level = 1
                if parent_id:
                    parent_node_result = await db.execute(
                        select(WorkflowNode).where(WorkflowNode.id == parent_id)
                    )
                    parent_node = parent_node_result.scalar_one_or_none()
                    if parent_node:
                        level = parent_node.level + 1

                node = WorkflowNode(
                    template_id=template.id,
                    parent_id=parent_id,
                    code=row["code"],
                    name=row["name"],
                    level=level,
                    order=0,
                    planned_days=row["planned_days"],
                    per_household=row["per_household"],
                    field_so_vb=row["field_so_vb"],
                    field_ngay_vb=row["field_ngay_vb"],
                    field_loai_vb=row["field_loai_vb"],
                    field_gia_tri_trinh=row["field_gia_tri_trinh"],
                    field_gia_tri_duyet=row["field_gia_tri_duyet"],
                    field_ghi_chu=row["field_ghi_chu"],
                    require_scan=row["require_scan"],
                    legal_basis=row["legal_basis"],
                    org_in_charge=row["org_in_charge"],
                )
                db.add(node)
                await db.flush()
                inserted += 1
                new_node_map[row["code"]] = node.id
                global_code_map[row["code"]] = node.id
        except Exception:
            upsert_errors += 1

    await db.commit()

    return {"imported": inserted + updated, "skipped": upsert_errors}


# GET /workflow/import-template
@router.get("/import-template")
async def download_import_template(
    current_user: User = Depends(get_current_user),
):
    """Download blank Excel template for workflow import."""
    from openpyxl.styles import PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quy trinh"

    # Header row
    headers = ["STT", "Mã bước", "Tên bước", "Parent code", "Thời gian (ngày)"]
    ws.append(headers)

    # Style header
    header_fill = PatternFill(start_color="9B1B30", end_color="9B1B30", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Sample data rows
    samples = [
        (1, "B1", "Thông báo thu hồi đất", None, 30),
        (2, "B1.1", "Lập phương án bồi thường", "B1", 15),
        (3, "B1.2", "Niêm yết phương án", "B1", 5),
        (4, "B2", "Chi trả bồi thường", None, 20),
    ]
    for row in samples:
        ws.append(list(row))

    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="mau-import-quy-trinh.xlsx"'},
    )


# S2-BE-06 — GET /workflow/export-excel
@router.get("/export-excel")
async def export_workflow_excel(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles(RoleEnum.admin)),
):
    # Get active template
    tmpl_result = await db.execute(
        select(WorkflowTemplate).where(WorkflowTemplate.is_active == True)  # noqa: E712
    )
    template = tmpl_result.scalar_one_or_none()
    if template is None:
        raise HTTPException(status_code=404, detail="Không tìm thấy workflow template đang hoạt động")

    nodes_result = await db.execute(
        select(WorkflowNode).where(WorkflowNode.template_id == template.id)
    )
    all_nodes: List[WorkflowNode] = list(nodes_result.scalars().all())

    # Build id->code map for parent_code resolution
    id_to_code: Dict[uuid.UUID, str] = {n.id: (n.code or "") for n in all_nodes}

    # DFS ordering
    children_map: Dict[Optional[uuid.UUID], List[WorkflowNode]] = {}
    for node in all_nodes:
        parent = node.parent_id
        children_map.setdefault(parent, []).append(node)
    for key in children_map:
        children_map[key].sort(key=lambda n: n.order)

    dfs_order: List[WorkflowNode] = []

    def dfs(parent_id: Optional[uuid.UUID]) -> None:
        for child in children_map.get(parent_id, []):
            dfs_order.append(child)
            dfs(child.id)

    dfs(None)

    # Build xlsx
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quy trinh template"

    bold_font = Font(bold=True)
    ws.append(EXCEL_COLUMNS)
    for cell in ws[1]:
        cell.font = bold_font

    for node in dfs_order:
        parent_code = id_to_code.get(node.parent_id, "") if node.parent_id else ""
        ws.append([
            node.code or "",
            parent_code,
            node.name,
            node.planned_days,
            node.per_household,
            node.field_so_vb,
            node.field_ngay_vb,
            node.field_loai_vb,
            node.field_gia_tri_trinh,
            node.field_gia_tri_duyet,
            node.field_ghi_chu,
            node.require_scan,
            node.legal_basis or "",
            node.org_in_charge or "",
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"quy-trinh-template-{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""},
    )
