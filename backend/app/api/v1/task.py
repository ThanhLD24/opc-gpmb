from __future__ import annotations

import io
import os
import uuid
from datetime import datetime
from typing import Optional, List, Dict, Any

from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from sqlalchemy.orm import selectinload
from pydantic import BaseModel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ...db.session import get_db
from ...db.models import (
    HoSoGPMB,
    HoSoWorkflowNode,
    TaskInstance,
    TaskAttachment,
    TaskStatusEnum,
    NodeHouseholdScope,
    Ho,
    User,
    RoleEnum,
    WorkflowNode,
    WorkflowNodeDocument,
)
from ...core.config import settings
from ...services.task_service import (
    assign_households_to_node,
    remove_household_from_node,
    update_task_rollup,
)
from ..deps import get_current_user, require_roles

router = APIRouter()


# ── Schemas ──────────────────────────────────────────────────────────────────

class TaskStatusUpdate(BaseModel):
    status: str


class TaskFieldsUpdate(BaseModel):
    so_vb: Optional[str] = None
    ngay_vb: Optional[str] = None  # ISO date string
    loai_vb: Optional[str] = None
    gia_tri_trinh: Optional[float] = None
    gia_tri_duyet: Optional[float] = None
    ghi_chu: Optional[str] = None


class ScopeUpdate(BaseModel):
    ho_ids: List[str]
    action: str = "assign"  # "assign" or "remove"


def task_to_dict(task: TaskInstance) -> Dict[str, Any]:
    return {
        "id": str(task.id),
        "ho_so_id": str(task.ho_so_id),
        "workflow_node_id": str(task.workflow_node_id),
        "ho_id": str(task.ho_id) if task.ho_id else None,
        "status": task.status.value,
        "so_vb": task.so_vb,
        "ngay_vb": task.ngay_vb.isoformat() if task.ngay_vb else None,
        "loai_vb": task.loai_vb,
        "gia_tri_trinh": task.gia_tri_trinh,
        "gia_tri_duyet": task.gia_tri_duyet,
        "ghi_chu": task.ghi_chu,
        "scan_file_path": task.scan_file_path,
        "completed_at": task.completed_at.isoformat() if task.completed_at else None,
        "created_at": task.created_at.isoformat(),
        "updated_at": task.updated_at.isoformat(),
    }


def task_to_dict_enriched(task: TaskInstance) -> Dict[str, Any]:
    """Enriched task dict that includes HoSoWorkflowNode metadata.
    Requires task.workflow_node to be eagerly loaded via selectinload,
    and task.attachments + node.source_node.documents to be loaded.
    """
    node = task.workflow_node
    is_leaf = bool(node and not node.children)

    # Guide documents from the source WorkflowNode template
    node_documents: List[Dict] = []
    if node and node.source_node and hasattr(node.source_node, "documents"):
        node_documents = [
            {
                "id": str(d.id),
                "ten_tai_lieu": d.ten_tai_lieu,
                "url": f"/uploads/{d.file_path}",
                "uploaded_at": d.uploaded_at.isoformat(),
            }
            for d in node.source_node.documents
        ]

    # Attachments on this task instance
    attachments: List[Dict] = []
    if hasattr(task, "attachments") and task.attachments:
        attachments = [
            {
                "id": str(a.id),
                "ten_tai_lieu": a.ten_tai_lieu,
                "url": f"/uploads/{a.file_path}",
                "uploaded_at": a.uploaded_at.isoformat(),
            }
            for a in task.attachments
        ]

    return {
        # Task fields
        "id": str(task.id),
        "ho_so_id": str(task.ho_so_id),
        "workflow_node_id": str(task.workflow_node_id),
        "ho_id": str(task.ho_id) if task.ho_id else None,
        "status": task.status.value,
        "so_vb": task.so_vb,
        "ngay_vb": task.ngay_vb.isoformat() if task.ngay_vb else None,
        "loai_vb": task.loai_vb,
        "gia_tri_trinh": task.gia_tri_trinh,
        "gia_tri_duyet": task.gia_tri_duyet,
        "ghi_chu": task.ghi_chu,
        "file_scan_url": f"/uploads/{task.scan_file_path}" if task.scan_file_path else None,
        "completed_at": task.completed_at.isoformat() if task.completed_at else None,
        "created_at": task.created_at.isoformat(),
        "updated_at": task.updated_at.isoformat(),
        # Node metadata
        "name": node.name if node else None,
        "code": node.code if node else None,
        "level": node.level if node else None,
        "per_household": node.per_household if node else False,
        "require_scan": node.require_scan if node else False,
        "is_leaf": is_leaf if node else False,
        "field_so_vb": node.field_so_vb if node else False,
        "field_ngay_vb": node.field_ngay_vb if node else False,
        "field_loai_vb": node.field_loai_vb if node else False,
        "field_gia_tri_trinh": node.field_gia_tri_trinh if node else False,
        "field_gia_tri_duyet": node.field_gia_tri_duyet if node else False,
        "field_ghi_chu": node.field_ghi_chu if node else False,
        # Documents & attachments
        "node_documents": node_documents,
        "attachments": attachments,
    }


def node_to_dict_with_tasks(
    node: HoSoWorkflowNode, tasks: List[TaskInstance]
) -> Dict[str, Any]:
    node_tasks = [t for t in tasks if t.workflow_node_id == node.id]
    return {
        "id": str(node.id),
        "source_node_id": str(node.source_node_id) if node.source_node_id else None,
        "parent_id": str(node.parent_id) if node.parent_id else None,
        "code": node.code,
        "name": node.name,
        "level": node.level,
        "order": node.order,
        "planned_days": node.planned_days,
        "per_household": node.per_household,
        "require_scan": node.require_scan,
        "field_so_vb": node.field_so_vb,
        "field_ngay_vb": node.field_ngay_vb,
        "field_loai_vb": node.field_loai_vb,
        "field_gia_tri_trinh": node.field_gia_tri_trinh,
        "field_gia_tri_duyet": node.field_gia_tri_duyet,
        "field_ghi_chu": node.field_ghi_chu,
        "tasks": [task_to_dict(t) for t in node_tasks],
        "children": [],
    }


def build_task_tree(
    nodes: List[HoSoWorkflowNode], tasks: List[TaskInstance]
) -> List[Dict]:
    node_map: Dict[str, Dict] = {}
    for node in sorted(nodes, key=lambda n: n.order):
        node_map[str(node.id)] = node_to_dict_with_tasks(node, tasks)

    roots = []
    for node in sorted(nodes, key=lambda n: n.order):
        d = node_map[str(node.id)]
        if node.parent_id and str(node.parent_id) in node_map:
            node_map[str(node.parent_id)]["children"].append(d)
        else:
            roots.append(d)

    def strip_empty_children(node_dict: Dict) -> Dict:
        if not node_dict["children"]:
            node_dict["children"] = None  # leaf — frontend checks children is None
        else:
            node_dict["children"] = [strip_empty_children(c) for c in node_dict["children"]]
        return node_dict

    return [strip_empty_children(r) for r in roots]


async def _get_ho_so_or_404(ho_so_id: str, db: AsyncSession) -> HoSoGPMB:
    result = await db.execute(
        select(HoSoGPMB).where(HoSoGPMB.id == uuid.UUID(ho_so_id))
    )
    hs = result.scalar_one_or_none()
    if hs is None:
        raise HTTPException(status_code=404, detail="Ho so not found")
    return hs


# ── Endpoints ────────────────────────────────────────────────────────────────

@router.get("/{ho_so_id}/tasks")
async def list_tasks(
    ho_so_id: str,
    ho_id: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Return tasks as a tree. Optionally filter by ho_id."""
    await _get_ho_so_or_404(ho_so_id, db)

    nodes_result = await db.execute(
        select(HoSoWorkflowNode).where(
            HoSoWorkflowNode.ho_so_id == uuid.UUID(ho_so_id)
        )
    )
    nodes = list(nodes_result.scalars().all())

    tasks_q = select(TaskInstance).where(
        TaskInstance.ho_so_id == uuid.UUID(ho_so_id)
    )
    if ho_id:
        tasks_q = tasks_q.where(
            (TaskInstance.ho_id == uuid.UUID(ho_id)) | (TaskInstance.ho_id == None)  # noqa: E711
        )
    tasks_result = await db.execute(tasks_q)
    tasks = list(tasks_result.scalars().all())

    return {"tree": build_task_tree(nodes, tasks)}


@router.get("/{ho_so_id}/tasks/{task_id}")
async def get_task(
    ho_so_id: str,
    task_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    result = await db.execute(
        select(TaskInstance)
        .options(
            selectinload(TaskInstance.workflow_node)
            .selectinload(HoSoWorkflowNode.children),
            selectinload(TaskInstance.workflow_node)
            .selectinload(HoSoWorkflowNode.source_node)
            .selectinload(WorkflowNode.documents),
            selectinload(TaskInstance.attachments),
        )
        .where(
            TaskInstance.id == uuid.UUID(task_id),
            TaskInstance.ho_so_id == uuid.UUID(ho_so_id),
        )
    )
    task = result.scalar_one_or_none()
    if task is None:
        raise HTTPException(status_code=404, detail="Task not found")
    return task_to_dict_enriched(task)


@router.patch("/{ho_so_id}/tasks/{task_id}/status")
async def update_task_status(
    ho_so_id: str,
    task_id: str,
    body: TaskStatusUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    # [SECURITY] Only admin/cbcq can change task completion status
    if current_user.role not in (RoleEnum.admin, RoleEnum.cbcq):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Chỉ admin hoặc CBCQ mới được cập nhật trạng thái công việc",
        )

    result = await db.execute(
        select(TaskInstance)
        .options(
            selectinload(TaskInstance.workflow_node)
            .selectinload(HoSoWorkflowNode.children),
            selectinload(TaskInstance.workflow_node)
            .selectinload(HoSoWorkflowNode.source_node)
            .selectinload(WorkflowNode.documents),
            selectinload(TaskInstance.attachments),
        )
        .where(
            TaskInstance.id == uuid.UUID(task_id),
            TaskInstance.ho_so_id == uuid.UUID(ho_so_id),
        )
    )
    task = result.scalar_one_or_none()
    if task is None:
        raise HTTPException(status_code=404, detail="Task not found")

    try:
        new_status = TaskStatusEnum(body.status)
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Invalid status: {body.status}")

    # Validate required fields before marking hoan_thanh
    if new_status == TaskStatusEnum.hoan_thanh:
        node = task.workflow_node
        if node:
            missing = []
            if node.field_so_vb and not task.so_vb:
                missing.append("Số văn bản")
            if node.field_ngay_vb and not task.ngay_vb:
                missing.append("Ngày văn bản")
            if node.field_loai_vb and not task.loai_vb:
                missing.append("Loại văn bản")
            if node.field_gia_tri_trinh and task.gia_tri_trinh is None:
                missing.append("Giá trị trình")
            if node.field_gia_tri_duyet and task.gia_tri_duyet is None:
                missing.append("Giá trị duyệt")
            if missing:
                raise HTTPException(
                    status_code=422,
                    detail=f"Chưa điền đủ thông tin bắt buộc: {', '.join(missing)}",
                )

    task.status = new_status
    if new_status == TaskStatusEnum.hoan_thanh:
        task.completed_at = datetime.utcnow()
    else:
        task.completed_at = None

    await db.flush()

    # Rollup
    await update_task_rollup(
        task.workflow_node_id, task.ho_id, uuid.UUID(ho_so_id), db
    )
    await db.commit()
    await db.refresh(task)
    # Re-fetch with node loaded to return enriched response
    result2 = await db.execute(
        select(TaskInstance)
        .options(
            selectinload(TaskInstance.workflow_node)
            .selectinload(HoSoWorkflowNode.children),
            selectinload(TaskInstance.workflow_node)
            .selectinload(HoSoWorkflowNode.source_node)
            .selectinload(WorkflowNode.documents),
            selectinload(TaskInstance.attachments),
        )
        .where(TaskInstance.id == task.id)
    )
    task = result2.scalar_one()
    return task_to_dict_enriched(task)


@router.patch("/{ho_so_id}/tasks/{task_id}/fields")
async def update_task_fields(
    ho_so_id: str,
    task_id: str,
    body: TaskFieldsUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    # [SECURITY] ke_toan can only update financial fields
    is_ke_toan = current_user.role == RoleEnum.ke_toan
    can_edit_all = current_user.role in (RoleEnum.admin, RoleEnum.cbcq)

    if not can_edit_all and not is_ke_toan:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Không có quyền cập nhật thông tin công việc",
        )

    result = await db.execute(
        select(TaskInstance)
        .options(
            selectinload(TaskInstance.workflow_node)
            .selectinload(HoSoWorkflowNode.children),
            selectinload(TaskInstance.workflow_node)
            .selectinload(HoSoWorkflowNode.source_node)
            .selectinload(WorkflowNode.documents),
            selectinload(TaskInstance.attachments),
        )
        .where(
            TaskInstance.id == uuid.UUID(task_id),
            TaskInstance.ho_so_id == uuid.UUID(ho_so_id),
        )
    )
    task = result.scalar_one_or_none()
    if task is None:
        raise HTTPException(status_code=404, detail="Task not found")

    # Administrative fields — chỉ admin/cbcq
    if can_edit_all:
        if body.so_vb is not None:
            task.so_vb = body.so_vb
        if body.ngay_vb is not None:
            try:
                task.ngay_vb = datetime.fromisoformat(body.ngay_vb)
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid ngay_vb format")
        if body.loai_vb is not None:
            task.loai_vb = body.loai_vb

    # Finance fields — admin/cbcq + ke_toan đều cập nhật được
    if body.gia_tri_trinh is not None:
        task.gia_tri_trinh = body.gia_tri_trinh
    if body.gia_tri_duyet is not None:
        task.gia_tri_duyet = body.gia_tri_duyet
    if body.ghi_chu is not None:
        task.ghi_chu = body.ghi_chu

    await db.commit()
    # Re-fetch with node loaded to return enriched response
    result2 = await db.execute(
        select(TaskInstance)
        .options(
            selectinload(TaskInstance.workflow_node)
            .selectinload(HoSoWorkflowNode.children),
            selectinload(TaskInstance.workflow_node)
            .selectinload(HoSoWorkflowNode.source_node)
            .selectinload(WorkflowNode.documents),
            selectinload(TaskInstance.attachments),
        )
        .where(TaskInstance.id == task.id)
    )
    task = result2.scalar_one()
    return task_to_dict_enriched(task)


@router.post("/{ho_so_id}/tasks/{task_id}/upload")
async def upload_scan_file(
    ho_so_id: str,
    task_id: str,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    result = await db.execute(
        select(TaskInstance).where(
            TaskInstance.id == uuid.UUID(task_id),
            TaskInstance.ho_so_id == uuid.UUID(ho_so_id),
        )
    )
    task = result.scalar_one_or_none()
    if task is None:
        raise HTTPException(status_code=404, detail="Task not found")

    # Save file
    upload_dir = os.path.join(settings.UPLOAD_DIR, ho_so_id)
    os.makedirs(upload_dir, exist_ok=True)
    filename = f"{task_id}_{file.filename}"
    file_path = os.path.join(upload_dir, filename)
    content = await file.read()
    with open(file_path, "wb") as f:
        f.write(content)

    url_path = f"/uploads/{ho_so_id}/{filename}"
    task.scan_file_path = url_path
    await db.commit()
    await db.refresh(task)
    return {"url": url_path, "task": task_to_dict(task)}


@router.get("/{ho_so_id}/nodes/{node_id}/scope")
async def get_node_scope(
    ho_so_id: str,
    node_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Get list of households assigned to this node."""
    result = await db.execute(
        select(NodeHouseholdScope).where(
            NodeHouseholdScope.workflow_node_id == uuid.UUID(node_id)
        )
    )
    scopes = result.scalars().all()

    ho_ids = [str(s.ho_id) for s in scopes]
    if ho_ids:
        ho_result = await db.execute(
            select(Ho).where(
                Ho.id.in_([s.ho_id for s in scopes])
            )
        )
        households = ho_result.scalars().all()
    else:
        households = []

    return {
        "node_id": node_id,
        "households": [
            {
                "id": str(h.id),
                "ma_ho": h.ma_ho,
                "ten_chu_ho": h.ten_chu_ho,
                "status": h.status.value,
            }
            for h in households
        ],
        "count": len(households),
    }


@router.put("/{ho_so_id}/nodes/{node_id}/scope")
async def update_node_scope(
    ho_so_id: str,
    node_id: str,
    body: ScopeUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles(RoleEnum.admin, RoleEnum.cbcq)),
):
    ho_ids = [uuid.UUID(hid) for hid in body.ho_ids]

    if body.action == "assign":
        result = await assign_households_to_node(
            uuid.UUID(ho_so_id), uuid.UUID(node_id), ho_ids, db
        )
        await db.commit()
        return {"action": "assign", **result}
    elif body.action == "remove":
        for ho_id in ho_ids:
            await remove_household_from_node(
                uuid.UUID(ho_so_id), uuid.UUID(node_id), ho_id, db
            )
        await db.commit()
        return {"action": "remove", "removed": len(ho_ids)}
    else:
        raise HTTPException(
            status_code=400, detail="action must be 'assign' or 'remove'"
        )


@router.get("/{ho_so_id}/pivot")
async def get_pivot(
    ho_so_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Return pivot matrix: rows=ho, columns=tasks (codes), values=status."""
    await _get_ho_so_or_404(ho_so_id, db)

    # Get all ho in this ho_so
    ho_result = await db.execute(
        select(Ho).where(Ho.ho_so_id == uuid.UUID(ho_so_id)).order_by(Ho.ma_ho)
    )
    households = list(ho_result.scalars().all())

    # Get all per_household nodes
    nodes_result = await db.execute(
        select(HoSoWorkflowNode).where(
            HoSoWorkflowNode.ho_so_id == uuid.UUID(ho_so_id),
            HoSoWorkflowNode.per_household == True,  # noqa: E712
            HoSoWorkflowNode.code != None,  # noqa: E711
        ).order_by(HoSoWorkflowNode.order)
    )
    nodes = list(nodes_result.scalars().all())

    # Get all tasks for per_household nodes
    tasks_result = await db.execute(
        select(TaskInstance).where(
            TaskInstance.ho_so_id == uuid.UUID(ho_so_id),
            TaskInstance.ho_id != None,  # noqa: E711
        )
    )
    tasks = list(tasks_result.scalars().all())

    # Build lookup: (ho_id, node_id) -> status
    task_map: Dict = {}
    for task in tasks:
        task_map[(str(task.ho_id), str(task.workflow_node_id))] = task.status.value

    columns = [
        {"node_id": str(n.id), "code": n.code, "name": n.name} for n in nodes
    ]
    rows = []
    for ho in households:
        row_data = {
            "ho_id": str(ho.id),
            "ma_ho": ho.ma_ho,
            "ten_chu_ho": ho.ten_chu_ho,
            "ho_status": ho.status.value,
            "tasks": {},
        }
        for node in nodes:
            key = (str(ho.id), str(node.id))
            row_data["tasks"][node.code] = task_map.get(key, None)
        rows.append(row_data)

    return {
        "columns": columns,
        "rows": rows,
        "total_ho": len(households),
        "total_nodes": len(nodes),
    }


@router.get("/{ho_so_id}/pivot/export")
async def export_pivot(
    ho_so_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Export pivot matrix as xlsx."""
    await _get_ho_so_or_404(ho_so_id, db)

    ho_result = await db.execute(
        select(Ho).where(Ho.ho_so_id == uuid.UUID(ho_so_id)).order_by(Ho.ma_ho)
    )
    households = list(ho_result.scalars().all())

    nodes_result = await db.execute(
        select(HoSoWorkflowNode).where(
            HoSoWorkflowNode.ho_so_id == uuid.UUID(ho_so_id),
            HoSoWorkflowNode.per_household == True,  # noqa: E712
            HoSoWorkflowNode.code != None,  # noqa: E711
        ).order_by(HoSoWorkflowNode.order)
    )
    nodes = list(nodes_result.scalars().all())

    tasks_result = await db.execute(
        select(TaskInstance).where(
            TaskInstance.ho_so_id == uuid.UUID(ho_so_id),
            TaskInstance.ho_id != None,  # noqa: E711
        )
    )
    tasks = list(tasks_result.scalars().all())

    task_map: Dict = {}
    for task in tasks:
        task_map[(str(task.ho_id), str(task.workflow_node_id))] = task.status.value

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pivot"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    done_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    pending_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Headers
    ws.cell(row=1, column=1, value="STT").fill = header_fill
    ws.cell(row=1, column=1).font = header_font
    ws.cell(row=1, column=2, value="Mã hộ").fill = header_fill
    ws.cell(row=1, column=2).font = header_font
    ws.cell(row=1, column=3, value="Tên chủ hộ").fill = header_fill
    ws.cell(row=1, column=3).font = header_font
    ws.cell(row=1, column=4, value="Trạng thái hộ").fill = header_fill
    ws.cell(row=1, column=4).font = header_font

    for col_idx, node in enumerate(nodes, start=5):
        cell = ws.cell(row=1, column=col_idx, value=node.code or node.name[:20])
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Data rows
    for row_idx, ho in enumerate(households, start=2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        ws.cell(row=row_idx, column=2, value=ho.ma_ho)
        ws.cell(row=row_idx, column=3, value=ho.ten_chu_ho)
        ws.cell(row=row_idx, column=4, value=ho.status.value)

        for col_idx, node in enumerate(nodes, start=5):
            key = (str(ho.id), str(node.id))
            task_status = task_map.get(key)
            cell = ws.cell(row=row_idx, column=col_idx)
            if task_status == "hoan_thanh":
                cell.value = "✓"
                cell.fill = done_fill
            elif task_status == "dang_thuc_hien":
                cell.value = "..."
                cell.fill = pending_fill
            else:
                cell.value = ""
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

    # Set column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 15
    for col_idx in range(5, 5 + len(nodes)):
        ws.column_dimensions[get_column_letter(col_idx)].width = 8

    ws.row_dimensions[1].height = 40

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=pivot_{ho_so_id}.xlsx"},
    )


# ── Task Attachments ─────────────────────────────────────────────────────────

@router.post("/{ho_so_id}/tasks/{task_id}/attachments", status_code=status.HTTP_201_CREATED)
async def upload_task_attachment(
    ho_so_id: str,
    task_id: str,
    ten_tai_lieu: str = Form(...),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role not in (RoleEnum.admin, RoleEnum.cbcq):
        raise HTTPException(status_code=403, detail="Không có quyền đính kèm văn bản")

    result = await db.execute(
        select(TaskInstance).where(
            TaskInstance.id == uuid.UUID(task_id),
            TaskInstance.ho_so_id == uuid.UUID(ho_so_id),
        )
    )
    task = result.scalar_one_or_none()
    if task is None:
        raise HTTPException(status_code=404, detail="Task not found")

    dest_dir = os.path.join(settings.UPLOAD_DIR, "task-attachments", task_id)
    os.makedirs(dest_dir, exist_ok=True)
    safe_name = f"{uuid.uuid4()}_{file.filename}"
    dest_path = os.path.join(dest_dir, safe_name)
    contents = await file.read()
    if len(contents) > 50 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="File quá lớn (tối đa 50MB)")
    with open(dest_path, "wb") as f:
        f.write(contents)

    relative_path = os.path.join("task-attachments", task_id, safe_name)
    attachment = TaskAttachment(
        task_instance_id=uuid.UUID(task_id),
        ten_tai_lieu=ten_tai_lieu,
        file_path=relative_path,
        uploaded_by=current_user.id,
    )
    db.add(attachment)
    await db.commit()
    await db.refresh(attachment)
    return {
        "id": str(attachment.id),
        "ten_tai_lieu": attachment.ten_tai_lieu,
        "url": f"/uploads/{attachment.file_path}",
        "uploaded_at": attachment.uploaded_at.isoformat(),
    }


@router.delete("/{ho_so_id}/tasks/{task_id}/attachments/{att_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_task_attachment(
    ho_so_id: str,
    task_id: str,
    att_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role not in (RoleEnum.admin, RoleEnum.cbcq):
        raise HTTPException(status_code=403, detail="Không có quyền xóa văn bản")

    result = await db.execute(
        select(TaskAttachment).where(
            TaskAttachment.id == uuid.UUID(att_id),
            TaskAttachment.task_instance_id == uuid.UUID(task_id),
        )
    )
    att = result.scalar_one_or_none()
    if att is None:
        raise HTTPException(status_code=404, detail="Attachment not found")

    try:
        full_path = os.path.join(settings.UPLOAD_DIR, att.file_path)
        if os.path.exists(full_path):
            os.remove(full_path)
    except OSError:
        pass

    await db.delete(att)
    await db.commit()
