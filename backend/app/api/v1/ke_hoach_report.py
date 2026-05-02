"""Báo cáo kế hoạch tháng — tổng hợp cross-ho-so cho CBCQ/Admin."""
from __future__ import annotations

import io
import uuid
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_
from sqlalchemy.orm import selectinload

from ...db.session import get_db
from ...db.models import HoSoGPMB, KeHoachThang, KeHoachThangItem, User, RoleEnum
from ..deps import get_current_user

router = APIRouter()

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


async def _get_allowed_ho_so(
    db: AsyncSession, current_user: User
) -> dict[uuid.UUID, HoSoGPMB]:
    """Return dict {ho_so.id: ho_so} for ho-so the user can see."""
    conditions = [HoSoGPMB.deleted_at.is_(None)]
    if current_user.role == RoleEnum.cbcq:
        conditions.append(HoSoGPMB.cbcq_id == current_user.id)
    result = await db.execute(select(HoSoGPMB).where(and_(*conditions)))
    return {hs.id: hs for hs in result.scalars().all()}


async def _load_ke_hoach(
    db: AsyncSession,
    thang: int,
    nam: int,
    allowed_ids: list[uuid.UUID],
) -> list[KeHoachThang]:
    if not allowed_ids:
        return []
    q = (
        select(KeHoachThang)
        .options(selectinload(KeHoachThang.items))
        .where(
            KeHoachThang.thang == thang,
            KeHoachThang.nam == nam,
            KeHoachThang.ho_so_id.in_(allowed_ids),
        )
        .order_by(KeHoachThang.ho_so_id)
    )
    result = await db.execute(q)
    return result.scalars().all()


@router.get("")
async def list_ke_hoach_report(
    thang: int = Query(..., ge=1, le=12),
    nam: int = Query(..., ge=2020),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Tổng hợp flat-list kế hoạch tháng cho tất cả hồ sơ user phụ trách."""
    ho_so_map = await _get_allowed_ho_so(db, current_user)
    ke_hoach_list = await _load_ke_hoach(db, thang, nam, list(ho_so_map.keys()))

    items_flat = []
    for kh in ke_hoach_list:
        hs = ho_so_map.get(kh.ho_so_id)
        if not hs:
            continue
        for item in sorted(kh.items, key=lambda i: i.thu_tu):
            items_flat.append({
                "ho_so_id": str(hs.id),
                "ho_so_code": hs.code,
                "ho_so_name": hs.name,
                "ke_hoach_id": str(kh.id),
                "ten_cong_viec": item.ten_cong_viec,
                "loai": "phat_sinh" if item.la_viec_phat_sinh else "quy_trinh",
                "ngay_du_kien": item.ngay_du_kien.isoformat() if item.ngay_du_kien else None,
                "da_hoan_thanh": item.da_hoan_thanh,
                "ghi_chu": item.ghi_chu,
            })

    return {
        "thang": thang,
        "nam": nam,
        "total_items": len(items_flat),
        "items": items_flat,
    }


@router.get("/export")
async def export_ke_hoach_report_excel(
    thang: int = Query(..., ge=1, le=12),
    nam: int = Query(..., ge=2020),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Xuất Excel tổng hợp kế hoạch tháng — tất cả hồ sơ user phụ trách."""
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(status_code=500, detail="openpyxl not available")

    ho_so_map = await _get_allowed_ho_so(db, current_user)
    ke_hoach_list = await _load_ke_hoach(db, thang, nam, list(ho_so_map.keys()))

    if not ke_hoach_list:
        raise HTTPException(
            status_code=404,
            detail=f"Không có kế hoạch tháng {thang}/{nam}",
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"KH {thang:02d}-{nam}"

    RED = "9B1B30"
    header_fill = PatternFill(start_color=RED, end_color=RED, fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    ho_so_fill = PatternFill(start_color="F5E6E8", end_color="F5E6E8", fill_type="solid")
    ho_so_font = Font(bold=True, color=RED, size=10)
    center_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = f"BÁO CÁO KẾ HOẠCH THÁNG {thang:02d}/{nam}"
    title_cell.font = Font(bold=True, color=RED, size=14)
    title_cell.alignment = center_align
    ws.row_dimensions[1].height = 28

    # Header
    headers = ["STT", "Hồ sơ", "Tên công việc", "Loại", "Ngày dự kiến", "Trạng thái", "Ghi chú"]
    ws.append(headers)
    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = cell_border
    ws.row_dimensions[2].height = 22

    col_widths = [6, 22, 40, 12, 16, 16, 30]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    row_num = 3
    stt = 1
    for kh in ke_hoach_list:
        hs = ho_so_map.get(kh.ho_so_id)
        if not hs:
            continue
        sorted_items = sorted(kh.items, key=lambda i: i.thu_tu)
        if not sorted_items:
            continue

        # Ho-so sub-header row
        ws.merge_cells(f"A{row_num}:G{row_num}")
        hs_cell = ws.cell(row=row_num, column=1)
        hs_cell.value = f"{hs.code}  —  {hs.name}"
        hs_cell.fill = ho_so_fill
        hs_cell.font = ho_so_font
        hs_cell.alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[row_num].height = 18
        row_num += 1

        for item in sorted_items:
            ngay_str = item.ngay_du_kien.strftime("%d/%m/%Y") if item.ngay_du_kien else ""
            loai = "Phát sinh" if item.la_viec_phat_sinh else "Quy trình"
            trang_thai = "Hoàn thành" if item.da_hoan_thanh else "Đang thực hiện"

            ws.append([stt, hs.code, item.ten_cong_viec, loai, ngay_str, trang_thai, item.ghi_chu or ""])
            for col_idx, cell in enumerate(ws[row_num], start=1):
                cell.border = cell_border
                cell.alignment = Alignment(vertical="center", wrap_text=(col_idx in (3, 7)))
            ws.cell(row=row_num, column=1).alignment = center_align
            row_num += 1
            stt += 1

    ws.append([])
    footer = ws.cell(row=row_num + 1, column=1)
    footer.value = f"Xuất ngày: {datetime.utcnow().strftime('%d/%m/%Y %H:%M')} UTC"
    footer.font = Font(italic=True, color="888888", size=9)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"ke-hoach-thang-{thang:02d}-{nam}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
