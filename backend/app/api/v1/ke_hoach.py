from __future__ import annotations

import io
import uuid
from datetime import datetime, date
from typing import Optional, List, Any, Dict

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_
from sqlalchemy.orm import selectinload
from pydantic import BaseModel, Field

from ...db.session import get_db
from ...db.models import (
    HoSoGPMB,
    KeHoachThang,
    KeHoachThangItem,
    TaskInstance,
    HoSoWorkflowNode,
    TaskStatusEnum,
    User,
    RoleEnum,
)
from ..deps import get_current_user, require_roles

router = APIRouter()

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Color
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


# ── Schemas ───────────────────────────────────────────────────────────────────

class KeHoachThangCreate(BaseModel):
    thang: int = Field(ge=1, le=12)
    nam: int = Field(ge=2020)
    ghi_chu: Optional[str] = None


class KeHoachThangItemUpdate(BaseModel):
    ngay_du_kien: Optional[str] = None  # "YYYY-MM-DD"
    ghi_chu: Optional[str] = None
    da_hoan_thanh: Optional[bool] = None  # chỉ áp dụng cho việc phát sinh


class ViecPhatSinhCreate(BaseModel):
    ten_cong_viec: str
    mo_ta: Optional[str] = None
    ngay_du_kien: Optional[str] = None  # "YYYY-MM-DD"
    ghi_chu: Optional[str] = None


# ── Helpers ───────────────────────────────────────────────────────────────────

def _item_to_dict(item: KeHoachThangItem) -> Dict[str, Any]:
    return {
        "id": str(item.id),
        "ke_hoach_thang_id": str(item.ke_hoach_thang_id),
        "task_instance_id": str(item.task_instance_id) if item.task_instance_id else None,
        "ten_cong_viec": item.ten_cong_viec,
        "mo_ta": item.mo_ta,
        "ngay_du_kien": item.ngay_du_kien.isoformat() if item.ngay_du_kien else None,
        "ghi_chu": item.ghi_chu,
        "la_viec_phat_sinh": item.la_viec_phat_sinh,
        "da_hoan_thanh": item.da_hoan_thanh,
        "thu_tu": item.thu_tu,
    }


def _ke_hoach_to_dict(kh: KeHoachThang, include_items: bool = False) -> Dict[str, Any]:
    d: Dict[str, Any] = {
        "id": str(kh.id),
        "ho_so_id": str(kh.ho_so_id),
        "thang": kh.thang,
        "nam": kh.nam,
        "created_by": str(kh.created_by) if kh.created_by else None,
        "created_at": kh.created_at.isoformat(),
        "da_xuat_bao_cao": kh.da_xuat_bao_cao,
        "ngay_xuat": kh.ngay_xuat.isoformat() if kh.ngay_xuat else None,
        "ghi_chu": kh.ghi_chu,
    }
    if include_items:
        sorted_items = sorted(kh.items, key=lambda i: i.thu_tu)
        d["items"] = [_item_to_dict(i) for i in sorted_items]
    return d


async def _get_ho_so_or_404(ho_so_id: str, db: AsyncSession) -> HoSoGPMB:
    result = await db.execute(
        select(HoSoGPMB).where(HoSoGPMB.id == uuid.UUID(ho_so_id))
    )
    hs = result.scalar_one_or_none()
    if hs is None:
        raise HTTPException(status_code=404, detail="Ho so not found")
    return hs


async def _get_ke_hoach_or_404(
    ho_so_id: str, kh_id: str, db: AsyncSession
) -> KeHoachThang:
    result = await db.execute(
        select(KeHoachThang)
        .options(selectinload(KeHoachThang.items))
        .where(
            KeHoachThang.id == uuid.UUID(kh_id),
            KeHoachThang.ho_so_id == uuid.UUID(ho_so_id),
        )
    )
    kh = result.scalar_one_or_none()
    if kh is None:
        raise HTTPException(status_code=404, detail="Ke hoach thang not found")
    return kh


async def _get_item_or_404(
    kh_id: str, item_id: str, db: AsyncSession
) -> KeHoachThangItem:
    result = await db.execute(
        select(KeHoachThangItem).where(
            KeHoachThangItem.id == uuid.UUID(item_id),
            KeHoachThangItem.ke_hoach_thang_id == uuid.UUID(kh_id),
        )
    )
    item = result.scalar_one_or_none()
    if item is None:
        raise HTTPException(status_code=404, detail="Item not found")
    return item


# ── S5-BE-01 Endpoints ────────────────────────────────────────────────────────

@router.get("/{ho_so_id}/ke-hoach")
async def list_ke_hoach(
    ho_so_id: str,
    thang: Optional[int] = Query(None, ge=1, le=12),
    nam: Optional[int] = Query(None, ge=2020),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """List ke hoach thang for a ho so, optionally filtered by thang/nam."""
    await _get_ho_so_or_404(ho_so_id, db)

    conditions = [KeHoachThang.ho_so_id == uuid.UUID(ho_so_id)]
    if thang is not None:
        conditions.append(KeHoachThang.thang == thang)
    if nam is not None:
        conditions.append(KeHoachThang.nam == nam)

    q = (
        select(KeHoachThang)
        .options(selectinload(KeHoachThang.items))
        .where(and_(*conditions))
        .order_by(KeHoachThang.nam.desc(), KeHoachThang.thang.desc())
    )
    result = await db.execute(q)
    kh_list = result.scalars().all()
    return [_ke_hoach_to_dict(kh, include_items=True) for kh in kh_list]


@router.post("/{ho_so_id}/ke-hoach/generate", status_code=status.HTTP_201_CREATED)
async def generate_ke_hoach(
    ho_so_id: str,
    body: KeHoachThangCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles(RoleEnum.admin, RoleEnum.cbcq)),
):
    """Auto-generate ke hoach thang from task_instances (status != hoan_thanh).
    Returns 409 if (ho_so_id, thang, nam) already exists.
    """
    await _get_ho_so_or_404(ho_so_id, db)

    # Check for duplicate
    existing = await db.execute(
        select(KeHoachThang).where(
            KeHoachThang.ho_so_id == uuid.UUID(ho_so_id),
            KeHoachThang.thang == body.thang,
            KeHoachThang.nam == body.nam,
        )
    )
    if existing.scalar_one_or_none() is not None:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"Ke hoach thang {body.thang}/{body.nam} da ton tai cho ho so nay",
        )

    # Pull task_instances with status != hoan_thanh, join workflow_node for ordering
    task_q = (
        select(TaskInstance)
        .options(selectinload(TaskInstance.workflow_node))
        .where(
            TaskInstance.ho_so_id == uuid.UUID(ho_so_id),
            TaskInstance.status != TaskStatusEnum.hoan_thanh,
        )
    )
    task_result = await db.execute(task_q)
    tasks = task_result.scalars().all()

    # Sort by workflow_node.level ASC, workflow_node.order ASC
    tasks_sorted = sorted(
        tasks,
        key=lambda t: (
            t.workflow_node.level if t.workflow_node else 999,
            t.workflow_node.order if t.workflow_node else 999,
        ),
    )

    # Create ke_hoach_thang record
    ke_hoach = KeHoachThang(
        ho_so_id=uuid.UUID(ho_so_id),
        thang=body.thang,
        nam=body.nam,
        created_by=current_user.id,
        ghi_chu=body.ghi_chu,
    )
    db.add(ke_hoach)
    await db.flush()  # get ke_hoach.id

    # Bulk insert items
    for idx, task in enumerate(tasks_sorted):
        node_name = task.workflow_node.name if task.workflow_node else "Công việc không xác định"
        item = KeHoachThangItem(
            ke_hoach_thang_id=ke_hoach.id,
            task_instance_id=task.id,
            ten_cong_viec=node_name,
            la_viec_phat_sinh=False,
            thu_tu=idx,
        )
        db.add(item)

    await db.commit()

    # Reload with items
    result = await db.execute(
        select(KeHoachThang)
        .options(selectinload(KeHoachThang.items))
        .where(KeHoachThang.id == ke_hoach.id)
    )
    ke_hoach = result.scalar_one()
    return _ke_hoach_to_dict(ke_hoach, include_items=True)


@router.get("/{ho_so_id}/ke-hoach/{kh_id}")
async def get_ke_hoach(
    ho_so_id: str,
    kh_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Get ke hoach detail with items."""
    kh = await _get_ke_hoach_or_404(ho_so_id, kh_id, db)
    return _ke_hoach_to_dict(kh, include_items=True)


@router.patch("/{ho_so_id}/ke-hoach/{kh_id}/items/{item_id}")
async def update_ke_hoach_item(
    ho_so_id: str,
    kh_id: str,
    item_id: str,
    body: KeHoachThangItemUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles(RoleEnum.admin, RoleEnum.cbcq)),
):
    """Update ngay_du_kien and/or ghi_chu of any item (including auto-gen)."""
    # Verify ke_hoach belongs to ho_so
    await _get_ke_hoach_or_404(ho_so_id, kh_id, db)
    item = await _get_item_or_404(kh_id, item_id, db)

    if body.ngay_du_kien is not None:
        try:
            item.ngay_du_kien = date.fromisoformat(body.ngay_du_kien)
        except ValueError:
            raise HTTPException(
                status_code=422, detail="Invalid ngay_du_kien format. Expected YYYY-MM-DD."
            )
    if body.ghi_chu is not None:
        item.ghi_chu = body.ghi_chu
    if body.da_hoan_thanh is not None:
        if not item.la_viec_phat_sinh:
            raise HTTPException(
                status_code=400,
                detail="Chỉ có thể cập nhật trạng thái cho việc phát sinh, không phải công việc quy trình",
            )
        item.da_hoan_thanh = body.da_hoan_thanh

    await db.commit()
    await db.refresh(item)
    return _item_to_dict(item)


# ── S5-BE-02 Endpoints ────────────────────────────────────────────────────────

@router.post("/{ho_so_id}/ke-hoach/{kh_id}/items", status_code=status.HTTP_201_CREATED)
async def create_viec_phat_sinh(
    ho_so_id: str,
    kh_id: str,
    body: ViecPhatSinhCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles(RoleEnum.admin, RoleEnum.cbcq)),
):
    """Add a viec phat sinh (ad-hoc task) to ke hoach."""
    kh = await _get_ke_hoach_or_404(ho_so_id, kh_id, db)

    # Determine next thu_tu
    max_thu_tu = max((i.thu_tu for i in kh.items), default=-1)

    ngay_du_kien_parsed: Optional[date] = None
    if body.ngay_du_kien is not None:
        try:
            ngay_du_kien_parsed = date.fromisoformat(body.ngay_du_kien)
        except ValueError:
            raise HTTPException(
                status_code=422, detail="Invalid ngay_du_kien format. Expected YYYY-MM-DD."
            )

    item = KeHoachThangItem(
        ke_hoach_thang_id=kh.id,
        task_instance_id=None,
        ten_cong_viec=body.ten_cong_viec,
        mo_ta=body.mo_ta,
        ngay_du_kien=ngay_du_kien_parsed,
        ghi_chu=body.ghi_chu,
        la_viec_phat_sinh=True,
        thu_tu=max_thu_tu + 1,
    )
    db.add(item)
    await db.commit()
    await db.refresh(item)
    return _item_to_dict(item)


@router.delete(
    "/{ho_so_id}/ke-hoach/{kh_id}/items/{item_id}",
    status_code=status.HTTP_204_NO_CONTENT,
)
async def delete_viec_phat_sinh(
    ho_so_id: str,
    kh_id: str,
    item_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles(RoleEnum.admin, RoleEnum.cbcq)),
):
    """Delete a viec phat sinh item. Returns 400 if item is from auto-gen (la_viec_phat_sinh=False)."""
    await _get_ke_hoach_or_404(ho_so_id, kh_id, db)
    item = await _get_item_or_404(kh_id, item_id, db)

    if not item.la_viec_phat_sinh:
        raise HTTPException(
            status_code=400,
            detail="Không thể xóa công việc từ quy trình",
        )

    await db.delete(item)
    await db.commit()


@router.get("/{ho_so_id}/ke-hoach/{kh_id}/export")
async def export_ke_hoach_excel(
    ho_so_id: str,
    kh_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Export ke hoach thang to Excel. Sets da_xuat_bao_cao=True + ngay_xuat=now()."""
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(status_code=500, detail="openpyxl not available")

    kh = await _get_ke_hoach_or_404(ho_so_id, kh_id, db)

    # Sort items by thu_tu
    sorted_items = sorted(kh.items, key=lambda i: i.thu_tu)

    # Build Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Ke hoach {kh.thang:02d}-{kh.nam}"  # max 31 chars

    headers = ["STT", "Tên công việc", "Mô tả", "Ngày dự kiến", "Ghi chú", "Loại"]
    ws.append(headers)

    # Style header row: bold, fill #9B1B30, white font
    header_fill = PatternFill(start_color="9B1B30", end_color="9B1B30", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for idx, item in enumerate(sorted_items, start=1):
        ngay_str = (
            item.ngay_du_kien.strftime("%d/%m/%Y") if item.ngay_du_kien else ""
        )
        loai = "Phát sinh" if item.la_viec_phat_sinh else "Quy trình"
        ws.append([
            idx,
            item.ten_cong_viec,
            item.mo_ta or "",
            ngay_str,
            item.ghi_chu or "",
            loai,
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Mark as exported
    kh.da_xuat_bao_cao = True
    kh.ngay_xuat = datetime.utcnow()
    await db.commit()

    filename = f"ke-hoach-thang-{kh.thang:02d}-{kh.nam}.xlsx"
    resp_headers = {
        "Content-Disposition": f'attachment; filename="{filename}"'
    }
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=resp_headers,
    )


@router.get("/{ho_so_id}/ke-hoach/{kh_id}/export/pdf")
async def export_ke_hoach_pdf(
    ho_so_id: str,
    kh_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Export ke hoach thang to PDF using reportlab."""
    if not REPORTLAB_AVAILABLE:
        raise HTTPException(status_code=500, detail="reportlab not available")

    hs = await _get_ho_so_or_404(ho_so_id, db)
    kh = await _get_ke_hoach_or_404(ho_so_id, kh_id, db)

    sorted_items = sorted(kh.items, key=lambda i: i.thu_tu)

    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    normal_style = styles["Normal"]

    story = []

    # Title
    story.append(Paragraph(
        f"Ke hoach chi tra thang {kh.thang:02d}/{kh.nam}",
        title_style,
    ))
    story.append(Spacer(1, 0.3 * cm))

    # Subtitle
    story.append(Paragraph(
        f"Ho so: {hs.code} -- {hs.name}",
        normal_style,
    ))
    story.append(Spacer(1, 0.5 * cm))

    # Table data
    headers = ["STT", "Ten cong viec", "Mo ta", "Ngay du kien", "Ghi chu", "Loai"]
    table_data = [headers]

    for idx, item in enumerate(sorted_items, start=1):
        ngay_str = item.ngay_du_kien.strftime("%d/%m/%Y") if item.ngay_du_kien else ""
        loai = "Phat sinh" if item.la_viec_phat_sinh else "Quy trinh"
        table_data.append([
            str(idx),
            item.ten_cong_viec or "",
            item.mo_ta or "",
            ngay_str,
            item.ghi_chu or "",
            loai,
        ])

    # Column widths for landscape A4 (~27 cm usable)
    col_widths = [1.2 * cm, 7 * cm, 6 * cm, 3 * cm, 5 * cm, 2.5 * cm]

    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        # Header row
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#9B1B30")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        # Data rows
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("ALIGN", (0, 1), (0, -1), "CENTER"),  # STT centered
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
        # Grid
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        # Word wrap for data cells
        ("WORDWRAP", (0, 0), (-1, -1), True),
    ]))
    story.append(table)
    story.append(Spacer(1, 0.8 * cm))

    # Footer
    today_str = datetime.utcnow().strftime("%d/%m/%Y")
    story.append(Paragraph(
        f"Da xuat bao cao: ngay {today_str}",
        normal_style,
    ))

    doc.build(story)
    output.seek(0)

    filename = f"ke-hoach-{kh.thang:02d}-{kh.nam}.pdf"
    resp_headers = {
        "Content-Disposition": f'attachment; filename="{filename}"'
    }
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers=resp_headers,
    )
